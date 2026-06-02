# DNTesterStation_MODERN_FINAL.py
# MODERN FINAL
# - Keeps v5 stability (Admin promote/rename, role protection)
# - Adds TRC Auto/Manual + Browse + Open Folder + Logout (Ctrl+L)
# - Improves FR title match and grid-safe value extraction

import re
import sqlite3
import uuid
import time
import lzma
import subprocess
import secrets
import hashlib
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Tuple
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

import os
import json
import shutil
import platform
import logging

try:
    from PIL import Image, ImageTk
except Exception:
    Image = ImageTk = None

try:
    import psutil
except Exception:
    psutil = None

try:
    from pywinauto import Desktop
except Exception:
    Desktop = None
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import Font, Alignment
except Exception:
    Workbook = None
    load_workbook = None
    get_column_letter = None
    Table = None
    TableStyleInfo = None
    Font = None
    Alignment = None


# =========================
# CONFIG
# =========================
TRC_DIR = r"C:\Users\alaa.khashoun\OneDrive - Diebold Nixdorf\Devicetrace"

# -------------------------
# LOG-ONLY VS1/VS2 SETTINGS
# -------------------------
# Folder that contains jdd_*.log files (offline test folder now; set to live folder later)
JDD_LOG_DIR = r"C:\Users\alaa.khashoun\OneDrive - Diebold Nixdorf\Pulpit\log_analysis"
# Use logs only (no Flight Recorder UI)
USE_JDD_LOG_EXTRACTION = True
# If True: require a log newer than the moment you press Start VS1/VS2 (recommended for live).
# If False: parse the latest existing logs immediately (recommended for offline tests).
JDD_REQUIRE_NEW_LOG = False
JDD_LOG_TIMEOUT_SEC = 60
JDD_LOG_POLL_MS = 900
JDD_LOG_STABLE_PASSES = 2


QUALIFYING_EXE = r"C:\Users\alaa.khashoun\OneDrive - Diebold Nixdorf\Pulpit\QualifyingPlusRM3\QualifyingPlus.exe"
QUALIFYING_PROCESS = "QualifyingPlus"
AUTO_OPEN_QUALIFYING = True

# If True, minimize this app while reading from Flight Recorder (improves VS1/VS2 stability)
AUTO_MINIMIZE_DURING_READ = False


FR_TITLE_RE = r"(?i).*flightrecorder.*"   # robust

MIN_TRC_SIZE_BYTES = 1024
WAIT_AFTER_OPEN_TRC_SEC = 2.5

RETRY_UI_SEC = 0.4
RETRY_COUNT = 120

SYSINFO_TEXT = "System Info"
KEY_MAT = "tcnFru"
KEY_SN = "snoFru"

VS_BOARD_TEXT = {
    "VS1": "CRS_SAFE:String2_VS1_CONTROLLER_BOARD",
    "VS2": "CRS_SAFE:String2_VS2_CONTROLLER_BOARD",
}
VS_BOARD_TEXT_FALLBACK = {
    "VS1": "CRS_SAFE:String2_VS1_CONTROLLER BOARD",
    "VS2": "CRS_SAFE:String2_VS2_CONTROLLER BOARD",
}


# --- TRC extraction mapping (fast, UI-free) ---
# These are snoElectronic values from Flight Recorder clipboard.
# Update if your station shows different snoElectronic for VS1/VS2.
USE_TRC_EXTRACTION = False
VS1_SNO_ELEC_CANDIDATES = {"86LE330156"}
VS2_SNO_ELEC_CANDIDATES = {"59R0818108"}

TIME_RE = re.compile(r"\b\d{2}:\d{2}:\d{2}\b")

COLORS = {
    "bg": "#f5f7fb",
    "card": "#ffffff",
    "header": "#133a75",
    "header_text": "#ffffff",
    "muted": "#475467",
    "accent": "#2d6cdf",
    "pass": "#17803d",
    "rework": "#c73b34",
    "doa": "#7c3aed",
    "border": "#d0d5dd",
    "danger": "#b42318",
    "gray": "#6b7280",
}


# UI branding
LOGO_FILE = 'logo.jpg'  # put the logo next to this .py (jpg/png)
LOGO_PATH = r"C:\Users\alaa.khashoun\OneDrive - Diebold Nixdorf\logo.jpg"  # optional absolute logo path
# Security (login hardening)
LOCKOUT_MAX_ATTEMPTS = 5
LOCKOUT_SECONDS = 60
REMEMBER_LAST_USERNAME = False


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "DNTesterStation"
DB_PATH = DATA_DIR / "tester_station_results.db"

# =========================
# Settings (per-station)
# =========================
APP_NAME = "DNTesterStation"

def _get_appdata_dir() -> Path:
    # Per-user writable directory (best for EXE + multi-station)
    base = os.environ.get("LOCALAPPDATA") or os.environ.get("APPDATA")
    if base:
        return Path(base) / APP_NAME
    return DATA_DIR

SETTINGS_PATH = _get_appdata_dir() / "settings.json"


# =========================
# Logging
# =========================
LOGGER_NAME = "DNTesterStation"

def get_log_path() -> Path:
    try:
        p = _get_appdata_dir() / "logs"
        p.mkdir(parents=True, exist_ok=True)
        return p / "app.log"
    except Exception:
        return Path(__file__).resolve().parent / "app.log"

_logger = logging.getLogger(LOGGER_NAME)
if not _logger.handlers:
    try:
        _logger.setLevel(logging.INFO)
        fh = logging.FileHandler(str(get_log_path()), encoding='utf-8')
        fmt = logging.Formatter('%(asctime)s | %(levelname)s | %(message)s')
        fh.setFormatter(fmt)
        _logger.addHandler(fh)
    except Exception:
        pass

def load_settings() -> dict:
    try:
        if SETTINGS_PATH.is_file():
            return json.loads(SETTINGS_PATH.read_text(encoding="utf-8")) or {}
    except Exception:
        pass
    return {}

def save_settings(data: dict) -> None:
    try:
        SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
        SETTINGS_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")
    except Exception:
        pass

def sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    # Replace Windows-illegal filename characters
    name = re.sub(r'[\\/:*?"<>|]+', '_', name)
    name = re.sub(r'\s+', '_', name)
    return name[:80] or "report"

# =========================
# Helpers
# =========================

def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_process_started(exe_path: str, process_name: str, timeout_sec: int = 20):
    if not exe_path:
        return

    if psutil is None:
        try:
            subprocess.Popen(exe_path, shell=True)
        except Exception:
            pass
        return

    for p in psutil.process_iter(["name"]):
        try:
            name = p.info.get("name") or ""
            if process_name.lower() in name.lower():
                return
        except Exception:
            pass

    try:
        subprocess.Popen(exe_path, shell=True)
    except Exception:
        return

    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        for p in psutil.process_iter(["name"]):
            try:
                name = p.info.get("name") or ""
                if process_name.lower() in name.lower():
                    return
            except Exception:
                pass
        time.sleep(0.5)


def pick_latest_trc(trc_dir: str) -> str:
    p = Path(trc_dir)
    if not p.is_dir():
        raise RuntimeError(f"TRC folder not found: {trc_dir}")

    trcs = [x for x in p.iterdir() if x.is_file() and x.suffix.lower() == ".trc"]
    if not trcs:
        raise RuntimeError("No .TRC files found")

    trcs.sort(key=lambda x: x.stat().st_mtime, reverse=True)

    for f in trcs:
        try:
            if f.stat().st_size >= MIN_TRC_SIZE_BYTES:
                return str(f)
        except Exception:
            continue

    return str(trcs[0])


def open_trc_via_windows(trc_path: str):
    subprocess.Popen(trc_path, shell=True)


def safe_invoke(ctrl) -> bool:
    try:
        ctrl.iface_invoke.Invoke(); return True
    except Exception:
        pass
    try:
        ctrl.invoke(); return True
    except Exception:
        pass
    try:
        ctrl.select(); return True
    except Exception:
        pass
    try:
        ctrl.set_focus(); return True
    except Exception:
        pass
    try:
        ctrl.click_input(); return True
    except Exception:
        return False


def find_visible_fr_window(timeout_sec: int = 120):
    if Desktop is None:
        raise RuntimeError("pywinauto not installed. Run: pip install pywinauto")

    desk = Desktop(backend="uia")
    deadline = time.time() + timeout_sec

    while time.time() < deadline:
        # regex match
        try:
            wins = desk.windows(title_re=FR_TITLE_RE)
        except Exception:
            wins = []

        for w in wins:
            try:
                if w.is_visible():
                    try:
                        w.set_focus()
                    except Exception:
                        pass
                    return w
            except Exception:
                pass

        # fallback scan
        try:
            all_wins = desk.windows()
        except Exception:
            all_wins = []

        for w in all_wins:
            try:
                if not w.is_visible():
                    continue
                title = (w.window_text() or "").strip()
                if title and "flightrecorder" in title.lower():
                    try:
                        w.set_focus()
                    except Exception:
                        pass
                    return w
            except Exception:
                pass

        time.sleep(0.5)

    raise RuntimeError("Flight Recorder window not ready. If FR runs as Admin, run terminal as Admin.")


def find_control_by_text(win, text: str):
    for c in win.descendants():
        try:
            if (c.window_text() or "") == text:
                return c
        except Exception:
            pass
    return None


def find_control_contains(win, needle1: str, needle2: str | None = None):
    n1 = (needle1 or "").lower()
    n2 = (needle2 or "").lower() if needle2 else ""

    for c in win.descendants():
        try:
            t = (c.window_text() or "")
            if not t:
                continue
            tl = t.lower()
            if n1 and n1 not in tl:
                continue
            if n2 and n2 not in tl:
                continue
            return c
        except Exception:
            pass
    return None


def find_tree_like_item(win, exact_text: str):
    """Return the best candidate in the LEFT navigation tree.

    This fixes VS2 reading wrong module when the same text appears in the detail pane.
    We only accept classes that behave like navigation items and choose the left-most one.
    """
    allowed = {"TreeItem", "ListItem", "DataItem"}
    candidates = []

    for c in win.descendants():
        try:
            if (c.window_text() or "") != exact_text:
                continue
            cls = c.friendly_class_name()
            if cls not in allowed:
                continue
            try:
                r = c.rectangle()
                # left-most = navigation pane
                candidates.append((r.left, r.top, cls, c))
            except Exception:
                candidates.append((99999, 99999, cls, c))
        except Exception:
            pass

    if not candidates:
        return None

    candidates.sort(key=lambda x: (x[0], x[1]))
    return candidates[0][3]


def click_reload_if_present(win) -> bool:
    for c in win.descendants():
        try:
            t = c.window_text() or ""
            if "Reload the actually loaded FlightRecord" in t:
                return safe_invoke(c)
        except Exception:
            pass
    try:
        win.type_keys("^r")
        return True
    except Exception:
        return False


def select_latest_timestamp(win) -> str:
    ts_tab = find_control_by_text(win, "Timestamp") or find_control_contains(win, "timestamp")
    if ts_tab:
        safe_invoke(ts_tab)
        time.sleep(0.25)

    for _ in range(RETRY_COUNT):
        candidates = []
        for c in win.descendants():
            try:
                txt = (c.window_text() or "").strip()
                if not txt:
                    continue
                if not TIME_RE.search(txt):
                    continue
                try:
                    y = c.rectangle().bottom
                except Exception:
                    y = 0
                candidates.append((y, txt, c))
            except Exception:
                continue

        if candidates:
            candidates.sort(key=lambda x: x[0])
            _, txt, ctrl = candidates[-1]
            safe_invoke(ctrl)
            return txt

        try:
            win.type_keys("{END}")
        except Exception:
            pass
        if ts_tab:
            safe_invoke(ts_tab)
        time.sleep(RETRY_UI_SEC)

    raise RuntimeError("Timestamp not found (not ready)")


def select_vs_board(win, vs: str):
    target = VS_BOARD_TEXT.get(vs)
    fallback = VS_BOARD_TEXT_FALLBACK.get(vs)

    # Prefer left navigation tree item
    if target:
        ctrl = find_tree_like_item(win, target)
        if ctrl and safe_invoke(ctrl):
            return

    # fallback exact match
    for _ in range(RETRY_COUNT):
        ctrl = find_control_by_text(win, target) if target else None
        if ctrl is None and fallback:
            ctrl = find_tree_like_item(win, fallback) or find_control_by_text(win, fallback)
        if ctrl and safe_invoke(ctrl):
            return
        time.sleep(RETRY_UI_SEC)

    raise RuntimeError(f"VS board not found: {vs}")


def read_key_value(win, key: str) -> str:
    for c in win.descendants():
        try:
            if (c.window_text() or "") == key:
                parent = c.parent()
                found = False
                for d in parent.descendants():
                    if found:
                        t = (d.window_text() or "").strip()
                        if t and t != key:
                            return t
                    if d == c:
                        found = True
        except Exception:
            pass
    return ""


def clean_material(val: str) -> str:
    return val.strip().split()[0] if val else ""


# =========================
# JDD log extraction (VS1/VS2, LOG-ONLY)
# =========================

def _extract_balanced_block(text: str, start_idx: int):
    # Bracket-balanced content starting after '['
    depth = 1
    i = start_idx
    n = len(text)
    while i < n:
        ch = text[i]
        if ch == '[':
            depth += 1
        elif ch == ']':
            depth -= 1
            if depth == 0:
                return text[start_idx:i], i + 1
        i += 1
    return '', start_idx


def _iter_blocks(text: str, token: str):
    # Yield bracket-balanced blocks for token like 'EepromTraceData['
    pos = 0
    while True:
        idx = text.find(token, pos)
        if idx == -1:
            break
        start = idx + len(token)
        body, end = _extract_balanced_block(text, start)
        if body:
            yield body
            pos = end
        else:
            pos = start + 1


def _val_after_key(body: str, key: str) -> str:
    # Extract 'key = VALUE' (up to next comma) from a block
    idx = body.find(key)
    if idx == -1:
        return ''
    eq = body.find('=', idx)
    if eq == -1:
        return ''
    s = eq + 1
    while s < len(body) and body[s] == ' ':
        s += 1
    e = body.find(',', s)
    if e == -1:
        e = len(body)
    return body[s:e].strip()


def _serial_from(v: str) -> str:
    v = (v or '').strip()
    if v.startswith('SerialNo['):
        j = v.find(']')
        return v[len('SerialNo['):j].strip() if j != -1 else ''
    return v


def _tcn_from(v: str) -> str:
    v = (v or '').strip()
    if v.startswith('TeamCenterNumber['):
        j = v.find(']')
        raw = v[len('TeamCenterNumber['):j].strip() if j != -1 else ''
    else:
        raw = v
    raw = (raw or '').strip()
    if not raw:
        return ''
    first = raw.split()[0]  # drop 'Rev X'
    if len(first) == 10 and first.startswith('1750'):
        first = '0' + first
    return first


def _jdd_pick_latest_log(log_dir: str) -> str:
    """Pick newest jdd_*.log by modification time."""
    p = Path(log_dir)
    if not p.is_dir():
        return ''
    cand = []
    for f in p.iterdir():
        if f.is_file() and f.name.lower().startswith('jdd_') and f.suffix.lower() == '.log':
            try:
                st = f.stat()
                cand.append((st.st_mtime, st.st_size, f))
            except Exception:
                pass
    if not cand:
        return ''
    cand.sort(key=lambda x: x[0], reverse=True)
    return str(cand[0][2])

    p = Path(log_dir)
    if not p.is_dir():
        return ''
    cand = []
    for f in p.iterdir():
        if f.is_file() and f.name.lower().startswith('jdd_') and f.suffix.lower() == '.log':
            try:
                st = f.stat()
                cand.append((st.st_mtime, st.st_size, f))
            except Exception:
                pass
    if not cand:
        return ''
    cand.sort(key=lambda x: x[0], reverse=True)
    # prefer logs containing VS identity markers
    for _, _, f in cand[:12]:
        try:
            s = f.read_text(errors='ignore')
            if ('snoElectronic=SerialNo[86LE330156]' in s) or ('snoElectronic=SerialNo[59R0818108]' in s):
                return str(f)
        except Exception:
            pass
    return str(cand[0][2])




# --- Smart log picker: newest log that contains VS identity snapshot ---
_VS_TCN_FRU_NEEDLE = 'tcnFru=TeamCenterNumber[01750200435'  # matches even if 'Rev X' follows


def _count_needle_in_file(path: Path, needle: str, max_count: int = 3, chunk_size: int = 1024 * 1024) -> int:
    """Count occurrences of needle in a file by streaming chunks. Stops after max_count."""
    try:
        needle_b = needle.encode('utf-8', errors='ignore')
        if not needle_b:
            return 0
        count = 0
        tail = b''
        with path.open('rb') as f:
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                buf = tail + chunk
                count += buf.count(needle_b)
                if count >= max_count:
                    return count
                # keep overlap
                tail = buf[-(len(needle_b) - 1):] if len(needle_b) > 1 else b''
        return count
    except Exception:
        return 0


def _jdd_pick_best_log_for_vs_material(log_dir: str, baseline_mtime: float = 0.0, target_vs: str | None = None) -> str:
    """Pick the newest jdd_*.log that contains the requested VS or the best recent VS log.

    Why >=2: we expect VS1 and VS2.

    baseline_mtime:
      - 0.0 => consider all logs (offline mode)
      - >0.0 => consider only logs newer than baseline (live mode)
    """
    p = Path(log_dir)
    if not p.is_dir():
        return ''

    logs = []
    for f in p.iterdir():
        if f.is_file() and f.name.lower().startswith('jdd_') and f.suffix.lower() == '.log':
            try:
                st = f.stat()
                if st.st_mtime > baseline_mtime:
                    logs.append((st.st_mtime, st.st_size, f))
            except Exception:
                pass

    if not logs:
        return ''

    logs.sort(key=lambda x: x[0], reverse=True)

    if target_vs:
        for _, _, f in logs[:60]:
            try:
                vsmap = extract_vs_from_jdd_log_file(str(f))
                if target_vs in vsmap:
                    return str(f)
            except Exception:
                pass

    # scan newest logs and choose first that contains >=2 hits
    for _, _, f in logs[:40]:
        hits = _count_needle_in_file(f, _VS_TCN_FRU_NEEDLE, max_count=3)
        if hits >= 2:
            return str(f)

    # fallback: choose newest log that contains at least 1 hit
    for _, _, f in logs[:40]:
        hits = _count_needle_in_file(f, _VS_TCN_FRU_NEEDLE, max_count=1)
        if hits >= 1:
            return str(f)

    # final fallback: newest log
    return str(logs[0][2])
def _jdd_latest_log_mtime(log_dir: str) -> float:
    try:
        p = Path(log_dir)
        if not p.is_dir():
            return 0.0
        mt = 0.0
        for f in p.iterdir():
            if f.is_file() and f.name.lower().startswith('jdd_') and f.suffix.lower() == '.log':
                try:
                    mt = max(mt, f.stat().st_mtime)
                except Exception:
                    pass
        return mt
    except Exception:
        return 0.0


def _jdd_find_new_log(log_dir: str, baseline_mtime: float) -> str:
    # Newest jdd_*.log with mtime > baseline
    p = Path(log_dir)
    if not p.is_dir():
        return ''
    cand = []
    for f in p.iterdir():
        if f.is_file() and f.name.lower().startswith('jdd_') and f.suffix.lower() == '.log':
            try:
                st = f.stat()
                if st.st_mtime > baseline_mtime:
                    cand.append((st.st_mtime, st.st_size, f))
            except Exception:
                pass
    if not cand:
        return ''
    cand.sort(key=lambda x: x[0], reverse=True)
    return str(cand[0][2])


def extract_vs_from_jdd_log_file(log_path: str):
    """Extract VS1/VS2 (tcnFru + snoFru) from one jdd log.

    Auto-detect by VS controller-board material tcnFru=01750200435.
    If 2+ matching EEPROM blocks exist, assign VS1/VS2 by ascending objectId.

    Returns dict like: {'VS1': (tcnFru, snoFru), 'VS2': (tcnFru, snoFru)}
    """
    TARGET_TCN_FRU = '01750200435'

    try:
        content = Path(log_path).read_text(errors='ignore')
    except Exception:
        return {}

    tokens = ('EepromTraceData[', 'EepromTraceData [')
    rows = []

    for token in tokens:
        for body in _iter_blocks(content, token):
            obj = _val_after_key(body, 'objectId')
            sno_fru = _serial_from(_val_after_key(body, 'snoFru'))
            tcn_fru = _tcn_from(_val_after_key(body, 'tcnFru'))
            if not sno_fru or not tcn_fru:
                continue
            if tcn_fru != TARGET_TCN_FRU:
                continue
            try:
                obj_i = int(obj)
            except Exception:
                obj_i = 10**18
            rows.append((obj_i, tcn_fru, sno_fru))

    # de-dup
    seen = set()
    uniq = []
    for obj_i, tcn_fru, sno_fru in rows:
        k = (tcn_fru, sno_fru)
        if k in seen:
            continue
        seen.add(k)
        uniq.append((obj_i, tcn_fru, sno_fru))

    if len(uniq) < 1:
        return {}

    uniq.sort(key=lambda x: x[0])
    out = {'VS1': (uniq[0][1], uniq[0][2])}
    if len(uniq) >= 2:
        out['VS2'] = (uniq[1][1], uniq[1][2])
    return out

    # Return {'VS1': (tcnFru, snoFru), 'VS2': (tcnFru, snoFru)} from one jdd log
    try:
        content = Path(log_path).read_text(errors='ignore')
    except Exception:
        return {}

    vsmap = {}
    for token in ('EepromTraceData[', 'EepromTraceData [', 'BvmEepromTraceData['):
        for body in _iter_blocks(content, token):
            sno_e = _serial_from(_val_after_key(body, 'snoElectronic'))
            sno_fru = _serial_from(_val_after_key(body, 'snoFru'))
            tcn_fru = _tcn_from(_val_after_key(body, 'tcnFru'))
            if not (sno_e and sno_fru and tcn_fru):
                continue
            if sno_e in VS1_SNO_ELEC_CANDIDATES:
                vsmap['VS1'] = (tcn_fru, sno_fru)
            if sno_e in VS2_SNO_ELEC_CANDIDATES:
                vsmap['VS2'] = (tcn_fru, sno_fru)
            if 'VS1' in vsmap and 'VS2' in vsmap:
                return vsmap
    return vsmap



def parse_clipboard_kv_first(text: str) -> dict:
    """Parse tab-separated clipboard dump (key	val1	val2...).
    Returns dict mapping key -> first value column (val1).
    """
    out = {}
    if not text:
        return out


# =========================
# TRC extraction (UI-free fast path)
# =========================
_TC_STRING = 0x74  # Java serialization TC_STRING
_RE_SNO_FRU  = re.compile(r"\d{10}$")
_RE_TCN_ELEC = re.compile(r"017501\d{5}$")
_RE_TCN_FRU  = re.compile(r"017502\d{5}$")
_RE_SNO_ELEC = re.compile(r"86[A-Z0-9]{8}$")


def _iter_tc_strings(dec: bytes):
    """Yield (byte_offset, string) for Java-serialized TC_STRING tokens."""
    i = 0
    n = len(dec)
    while i < n - 3:
        if dec[i] == _TC_STRING:
            ln = (dec[i + 1] << 8) | dec[i + 2]
            if 0 < ln < 256 and i + 3 + ln <= n:
                s = dec[i + 3:i + 3 + ln]
                if all(32 <= b <= 126 for b in s):
                    yield i, s.decode("ascii", errors="ignore")
                i += 3 + ln
                continue
        i += 1


def _classify_cluster(strings):
    """Given a list of strings starting at a snoFru, attempt to extract fields."""
    sno_fru = strings[0]
    tcn_e = sno_e = tcn_fru = ""
    for s in strings[1:26]:
        if not tcn_e and _RE_TCN_ELEC.fullmatch(s):
            tcn_e = s
        elif not sno_e and _RE_SNO_ELEC.fullmatch(s):
            sno_e = s
        elif not tcn_fru and _RE_TCN_FRU.fullmatch(s):
            tcn_fru = s
        if tcn_e and sno_e and tcn_fru:
            break
    if not (tcn_e and sno_e and tcn_fru):
        return None
    return {"snoFru": sno_fru, "tcnElectronic": tcn_e, "snoElectronic": sno_e, "tcnFru": tcn_fru}


def extract_vs_from_trc(trc_path: str):
    """Extract VS1/VS2 (tcnFru + snoFru) directly from RM3 *.CRYPT.TRC.

    Returns dict: {"VS1": (tcnFru, snoFru), "VS2": (tcnFru, snoFru)}
    Uses snoElectronic anchors configured in VS1_SNO_ELEC_CANDIDATES / VS2_SNO_ELEC_CANDIDATES.
    """
    raw = Path(trc_path).read_bytes()
    dec = lzma.decompress(raw, format=lzma.FORMAT_ALONE)

    all_strings = [(off, s) for off, s in _iter_tc_strings(dec)]
    records = []
    for idx, (off, s) in enumerate(all_strings):
        if not _RE_SNO_FRU.fullmatch(s):
            continue
        win = [x[1] for x in all_strings[idx: idx + 60]]
        rec = _classify_cluster(win)
        if rec:
            rec["offset"] = off
            records.append(rec)

    if not records:
        raise RuntimeError("No VS records found in decompressed TRC")

    vs1_cand = [r for r in records if r.get('snoElectronic') in VS1_SNO_ELEC_CANDIDATES]
    vs2_cand = [r for r in records if r.get('snoElectronic') in VS2_SNO_ELEC_CANDIDATES]

    out = {}
    if vs1_cand:
        r = max(vs1_cand, key=lambda x: x['offset'])
        out['VS1'] = (r['tcnFru'], r['snoFru'])
    if vs2_cand:
        r = max(vs2_cand, key=lambda x: x['offset'])
        out['VS2'] = (r['tcnFru'], r['snoFru'])

    if 'VS1' not in out or 'VS2' not in out:
        uniq = sorted({r['snoElectronic'] for r in records})
        raise RuntimeError(f"Could not map VS1/VS2. Available snoElectronic: {uniq}")

    return out
    for ln in text.splitlines():
        ln = ln.strip()
        if not ln:
            continue
        parts = ln.split("	")
        if len(parts) < 2:
            continue
        k = parts[0].strip()
        v = parts[1].strip()
        if k and v:
            out[k] = v
    return out

# =========================
# Password hashing
# =========================

def hash_password(password: str, salt: bytes = None, iterations: int = 200_000):
    if salt is None:
        salt = secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return salt, dk, iterations


def verify_password(password: str, salt: bytes, dk: bytes, iterations: int) -> bool:
    test = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return secrets.compare_digest(test, dk)

# =========================
# Database
# =========================

class Database:
    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(self.path))
        self.conn.execute("PRAGMA journal_mode=WAL;")
        self.conn.execute("PRAGMA foreign_keys=ON;")
        self._init()

    def _init(self):
        cur = self.conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                role TEXT NOT NULL,
                salt BLOB NOT NULL,
                pw_hash BLOB NOT NULL,
                iterations INTEGER NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                session_id TEXT PRIMARY KEY,
                tester TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'tech',
                vs TEXT NOT NULL,
                material TEXT NOT NULL,
                serial TEXT NOT NULL,
                result TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                trc_path TEXT NOT NULL,
                trc_time TEXT DEFAULT ''
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS session_files (
                file_id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_path TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                file_data BLOB NOT NULL,
                saved_at TEXT NOT NULL
            )
        """)

        # Migration for older DBs
        cur.execute("PRAGMA table_info(sessions)")
        cols = {r[1] for r in cur.fetchall()}
        if "role" not in cols:
            cur.execute("ALTER TABLE sessions ADD COLUMN role TEXT NOT NULL DEFAULT 'tech'")
        if "trc_time" not in cols:
            cur.execute("ALTER TABLE sessions ADD COLUMN trc_time TEXT DEFAULT ''")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS auth_lockouts (
                username TEXT PRIMARY KEY,
                fail_count INTEGER NOT NULL DEFAULT 0,
                locked_until REAL NOT NULL DEFAULT 0
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                actor TEXT NOT NULL,
                action TEXT NOT NULL,
                details TEXT NOT NULL
            )
        """)

        self.conn.commit()

    # users
    def has_any_admin(self) -> bool:
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM users WHERE role='admin' AND active=1")
        return (cur.fetchone()[0] or 0) > 0

    def create_user(self, username: str, password: str, role: str):
        u = username.strip()
        if not u:
            raise ValueError("Username required")
        if role not in ("admin", "tech"):
            raise ValueError("Invalid role")
        salt, dk, iters = hash_password(password)
        self.conn.execute(
            "INSERT INTO users (username, role, salt, pw_hash, iterations, active, created_at) VALUES (?, ?, ?, ?, ?, 1, ?)",
            (u, role, salt, dk, iters, now_ts())
        )
        self.conn.commit()

    def authenticate(self, username: str, password: str):
        cur = self.conn.cursor()
        cur.execute("SELECT role, salt, pw_hash, iterations, active FROM users WHERE username=?", (username,))
        row = cur.fetchone()
        if not row:
            return None
        role, salt, pw_hash, iters, active = row
        if active != 1:
            return None
        return role if verify_password(password, salt, pw_hash, iters) else None

    def list_users(self):
        cur = self.conn.cursor()
        cur.execute("SELECT username, role, active, created_at FROM users ORDER BY role DESC, username ASC")
        return cur.fetchall()

    def set_user_active(self, username: str, active: bool):
        self.conn.execute("UPDATE users SET active=? WHERE username=?", (1 if active else 0, username))
        self.conn.commit()

    def reset_password(self, username: str, new_password: str):
        salt, dk, iters = hash_password(new_password)
        self.conn.execute("UPDATE users SET salt=?, pw_hash=?, iterations=? WHERE username=?", (salt, dk, iters, username))
        self.conn.commit()

    def set_role(self, username: str, role: str):
        if role not in ("admin", "tech"):
            raise ValueError("Invalid role")
        self.conn.execute("UPDATE users SET role=? WHERE username=?", (role, username))
        self.conn.commit()

    def rename_user(self, old_username: str, new_username: str):
        old_u = old_username.strip()
        new_u = new_username.strip()
        if not new_u:
            raise ValueError("New username required")
        # update users
        self.conn.execute("UPDATE users SET username=? WHERE username=?", (new_u, old_u))
        # also update sessions for history consistency
        self.conn.execute("UPDATE sessions SET tester=? WHERE tester=?", (new_u, old_u))
        self.conn.commit()

    # sessions
    def insert_session(self, tester, role, vs, material, serial, result, start_time, end_time, trc_path, trc_time):
        sid = str(uuid.uuid4())
        self.conn.execute(
            "INSERT INTO sessions (session_id, tester, role, vs, material, serial, result, start_time, end_time, trc_path, trc_time) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (sid, tester, role, vs, material, serial, result, start_time, end_time, trc_path, trc_time or "")
        )
        self.conn.commit()
        return sid
    def save_trc_blob(self, session_id: str, trc_path: str):
        p = Path(trc_path)
        data = p.read_bytes()

    def audit_event(self, actor: str, action: str, details: str = ""):
        try:
            cur = self.conn.cursor()
            cur.execute("INSERT INTO audit_log(ts, actor, action, details) VALUES(?,?,?,?)",
                        (now_ts(), actor or "", action or "", details or ""))
            self.conn.commit()
        except Exception:
            pass

    def lockout_status(self, username: str):
        """Return (is_locked: bool, seconds_left: int, fail_count: int)."""
        try:
            cur = self.conn.cursor()
            cur.execute("SELECT fail_count, locked_until FROM auth_lockouts WHERE username=?", (username,))
            row = cur.fetchone()
            if not row:
                return False, 0, 0
            fail_count, locked_until = row
            now = time.time()
            if locked_until and locked_until > now:
                return True, int(locked_until - now), int(fail_count or 0)
            return False, 0, int(fail_count or 0)
        except Exception:
            return False, 0, 0

    def record_login_attempt(self, username: str, success: bool):
        """Update lockout counter. On success resets. On fail increments and locks if needed."""
        if not username:
            return
        try:
            cur = self.conn.cursor()
            cur.execute("SELECT fail_count, locked_until FROM auth_lockouts WHERE username=?", (username,))
            row = cur.fetchone()
            fail_count = int(row[0]) if row else 0
            locked_until = float(row[1]) if row else 0.0
            now = time.time()
            if success:
                fail_count = 0
                locked_until = 0.0
            else:
                if locked_until and locked_until <= now:
                    locked_until = 0.0
                fail_count += 1
                if fail_count >= int(LOCKOUT_MAX_ATTEMPTS):
                    locked_until = now + float(LOCKOUT_SECONDS)
            cur.execute(
                "INSERT INTO auth_lockouts(username, fail_count, locked_until) VALUES(?,?,?) "
                "ON CONFLICT(username) DO UPDATE SET fail_count=excluded.fail_count, locked_until=excluded.locked_until",
                (username, fail_count, locked_until),
            )
            self.conn.commit()
        except Exception:
            pass

        self.conn.execute(
            "INSERT INTO session_files (session_id, file_name, file_path, file_size, file_data, saved_at) VALUES (?, ?, ?, ?, ?, ?)",
            (session_id, p.name, str(p), len(data), data, now_ts())
        )
        self.conn.commit()

    def update_session_result(self, session_id: str, new_result: str):
        self.conn.execute("UPDATE sessions SET result=? WHERE session_id=?", (new_result, session_id))
        self.conn.commit()



    def fetch_session(self, session_id: str):
        """Fetch one session row for export: (session_id, tester, role, vs, material, serial, result, start_time, end_time, trc_path, trc_time)."""
        cur = self.conn.cursor()
        cur.execute(
            "SELECT session_id, tester, role, vs, material, serial, result, start_time, end_time, trc_path, trc_time FROM sessions WHERE session_id=?",
            (session_id,)
        )
        return cur.fetchone()

    def fetch_recent(self, limit=200):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT session_id, tester, role, vs, material, serial, result, start_time, end_time, trc_path, trc_time FROM sessions ORDER BY end_time DESC LIMIT ?",
            (limit,)
        )
        return cur.fetchall()

# =========================
# UI dialogs
# =========================

@dataclass
class LoginResult:
    tester: str
    role: str

class CreateAdminDialog(tk.Toplevel):
    def __init__(self, master, db: Database):
        super().__init__(master)
        self.db = db
        self.created = False
        self.title("Create Admin")
        self.resizable(False, False)
        self.configure(bg=COLORS["bg"])
        self.protocol("WM_DELETE_WINDOW", self._close)

        self.u = tk.StringVar(); self.p1 = tk.StringVar(); self.p2 = tk.StringVar(); self.msg = tk.StringVar(value="")

        card = tk.Frame(self, bg=COLORS["card"], bd=1, relief="solid")
        card.pack(padx=14, pady=14, fill="both", expand=True)

        tk.Label(card, text="First Run: Create Admin", bg=COLORS["card"], font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=14, pady=(12, 6))

        frm = tk.Frame(card, bg=COLORS["card"])
        frm.pack(fill="x", padx=14)
        tk.Label(frm, text="Admin Username", bg=COLORS["card"], fg=COLORS["muted"]).grid(row=0, column=0, sticky="w")
        tk.Entry(frm, textvariable=self.u).grid(row=1, column=0, sticky="ew", pady=(2, 10))
        tk.Label(frm, text="Password", bg=COLORS["card"], fg=COLORS["muted"]).grid(row=2, column=0, sticky="w")
        tk.Entry(frm, textvariable=self.p1, show="*").grid(row=3, column=0, sticky="ew", pady=(2, 10))
        tk.Label(frm, text="Repeat Password", bg=COLORS["card"], fg=COLORS["muted"]).grid(row=4, column=0, sticky="w")
        tk.Entry(frm, textvariable=self.p2, show="*").grid(row=5, column=0, sticky="ew", pady=(2, 10))
        frm.grid_columnconfigure(0, weight=1)

        tk.Label(card, textvariable=self.msg, bg=COLORS["card"], fg=COLORS["danger"], font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=14, pady=(0, 8))
        tk.Button(card, text="Create Admin", bg=COLORS["accent"], fg="white", relief="flat", font=("Segoe UI", 11, "bold"), command=self._create).pack(fill="x", padx=14, pady=(0, 14))

        try:
            self.bind('<Return>', lambda e: self._create())
            self.bind('<Escape>', lambda e: self._close())
        except Exception:
            pass

        self.grab_set(); self.focus_force()

    def _create(self):
        u = self.u.get().strip(); p1 = self.p1.get(); p2 = self.p2.get()
        if not u or not p1:
            self.msg.set("Username and password required"); return
        if p1 != p2:
            self.msg.set("Passwords do not match"); return
        try:
            self.db.create_user(u, p1, role="admin")
        except Exception as e:
            self.msg.set(str(e)); return
        self.created = True
        self.destroy()

    def _close(self):
        self.created = False
        self.destroy()

class LoginDialog(tk.Toplevel):
    def __init__(self, master, db: Database):
        super().__init__(master)
        self.db = db
        self.result = None
        self.title("Login")
        self.resizable(True, True)
        self.configure(bg=COLORS["bg"])
        self.protocol("WM_DELETE_WINDOW", self._close)

        self.user = tk.StringVar(); self.pw = tk.StringVar(); self.guest = tk.StringVar(); self.msg = tk.StringVar(value="")

        try:
            self.minsize(560, 360)
        except Exception:
            pass

        root = tk.Frame(self, bg=COLORS["bg"])
        root.grid(row=0, column=0, sticky="nsew", padx=14, pady=14)
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        root.columnconfigure(0, weight=0)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(0, weight=1)

        banner_w = 220
        banner = tk.Frame(root, bg=COLORS["header"], width=banner_w)
        banner.grid(row=0, column=0, sticky="ns")
        banner.grid_propagate(False)

        # Logo in banner (branding)
        self._banner_logo_img = None
        if Image is not None:
            try:
                lp = Path(LOGO_PATH) if (globals().get('LOGO_PATH') and str(LOGO_PATH).strip()) else (Path(__file__).resolve().parent / LOGO_FILE)
                if lp.exists():
                    img = Image.open(str(lp))
                    max_w = banner_w - 28
                    w, h = img.size
                    scale = min(1.0, max_w / float(w))
                    img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
                    self._banner_logo_img = ImageTk.PhotoImage(img)
                    tk.Label(banner, image=self._banner_logo_img, bg=COLORS["header"]).pack(anchor="nw", padx=14, pady=(14, 6))
            except Exception:
                pass

        tk.Label(banner, text="DN Tester Station", bg=COLORS["header"], fg=COLORS["header_text"],
                 font=("Segoe UI", 15, "bold"), padx=14, pady=10, justify="left", wraplength=banner_w-28).pack(anchor="nw")
        tk.Label(banner, text="Secure sign-in\nLog-only VS extraction", bg=COLORS["header"], fg=COLORS["header_text"],
                 font=("Segoe UI", 10), padx=14, pady=6, justify="left", wraplength=banner_w-28).pack(anchor="nw")

        card = tk.Frame(root, bg=COLORS["card"], bd=1, relief="solid")
        card.grid(row=0, column=1, sticky="nsew")
        card.columnconfigure(0, weight=1)

        inner = tk.Frame(card, bg=COLORS["card"], padx=18, pady=16)
        inner.grid(row=0, column=0, sticky="nsew")
        inner.columnconfigure(0, weight=1)

        tk.Label(inner, text="Login", bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", pady=(0,10))

        form = tk.Frame(inner, bg=COLORS["card"])
        form.grid(row=1, column=0, sticky="ew")
        form.columnconfigure(0, weight=1)

        tk.Label(form, text="Username", bg=COLORS["card"], fg=COLORS["muted"]).grid(row=0, column=0, sticky="w")
        ent_user = ttk.Entry(form, textvariable=self.user)
        ent_user.grid(row=1, column=0, sticky="ew", pady=(2,10))

        tk.Label(form, text="Password", bg=COLORS["card"], fg=COLORS["muted"]).grid(row=2, column=0, sticky="w")
        ent_pw = ttk.Entry(form, textvariable=self.pw, show="•")
        ent_pw.grid(row=3, column=0, sticky="ew", pady=(2,12))

        btns = tk.Frame(form, bg=COLORS["card"])
        btns.grid(row=4, column=0, sticky="ew")
        btns.columnconfigure(0, weight=1); btns.columnconfigure(1, weight=1)
        ttk.Button(btns, text="Login", command=self._login).grid(row=0, column=0, sticky="ew", padx=(0,6))
        ttk.Button(btns, text="Guest", command=self._guest).grid(row=0, column=1, sticky="ew", padx=(6,0))

        sep = tk.Frame(form, bg=COLORS["border"], height=1)
        sep.grid(row=5, column=0, sticky="ew", pady=12)

        tk.Label(form, text="Guest name (required)", bg=COLORS["card"], fg=COLORS["muted"]).grid(row=6, column=0, sticky="w")
        ttk.Entry(form, textvariable=self.guest).grid(row=7, column=0, sticky="ew", pady=(2,6))

        tk.Label(form, textvariable=self.msg, bg=COLORS["card"], fg=COLORS["danger"], font=("Segoe UI",10,"bold")).grid(row=8, column=0, sticky="w", pady=(4,0))

        try:
            ent_user.focus_set()
            self.bind('<Return>', lambda e: self._login())
        except Exception:
            pass

        try:
            self.grab_set(); self.focus_force()
        except Exception:
            pass

    def _login(self):
        u = self.user.get().strip(); p = self.pw.get()
        if not u or not p:
            self.msg.set("Enter username and password"); return

        locked, secs_left, fail_count = self.db.lockout_status(u)
        if locked:
            self.msg.set(f"Account locked. Try again in {secs_left}s")
            self.db.audit_event(u, "login_blocked", f"seconds_left={secs_left} fail_count={fail_count}")
            return

        role = self.db.authenticate(u, p)
        if not role:
            self.db.record_login_attempt(u, success=False)
            locked2, secs_left2, fail_count2 = self.db.lockout_status(u)
            if locked2:
                self.msg.set(f"Too many attempts. Locked for {secs_left2}s")
                self.db.audit_event(u, "login_locked", f"fail_count={fail_count2}")
            else:
                self.msg.set("Invalid credentials")
                self.db.audit_event(u, "login_failed", f"fail_count={fail_count2}")
            return

        self.db.record_login_attempt(u, success=True)
        self.db.audit_event(u, "login_success", f"role={role}")
        self.result = LoginResult(tester=u, role=role)
        self.destroy()

    def _guest(self):
        n = self.guest.get().strip()
        if not n:
            self.msg.set("Guest name required"); return
        self.db.audit_event(f"Guest - {n}", "guest_login", "")
        self.result = LoginResult(tester=f"Guest - {n}", role="guest")
        self.destroy()

    def _close(self):
        self.result = None
        self.destroy()

class AdminPanel(tk.Toplevel):
    def __init__(self, master, db: Database):
        super().__init__(master)
        self.db = db
        self.title("Admin Panel")
        self.geometry("760x480")
        self.configure(bg=COLORS["bg"])

        self.var_user = tk.StringVar()
        self.var_pass = tk.StringVar()

        self._build()
        self._refresh()

    def _build(self):
        top = tk.Frame(self, bg=COLORS["card"], bd=1, relief="solid")
        top.pack(fill="x", padx=14, pady=14)

        tk.Label(top, text="Create User", bg=COLORS["card"], font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=12, pady=(10, 6))

        tk.Label(top, text="Username", bg=COLORS["card"], fg=COLORS["muted"]).grid(row=1, column=0, sticky="w", padx=12)
        tk.Entry(top, textvariable=self.var_user).grid(row=1, column=1, sticky="ew", padx=12)

        tk.Label(top, text="Password", bg=COLORS["card"], fg=COLORS["muted"]).grid(row=2, column=0, sticky="w", padx=12, pady=(6, 0))
        tk.Entry(top, textvariable=self.var_pass, show="*").grid(row=2, column=1, sticky="ew", padx=12, pady=(6, 0))

        tk.Button(top, text="Create Technician", bg=COLORS["accent"], fg="white", relief="flat", command=lambda: self._create_user("tech")).grid(row=3, column=0, sticky="ew", padx=12, pady=10)
        tk.Button(top, text="Create Admin", bg=COLORS["doa"], fg="white", relief="flat", command=lambda: self._create_user("admin")).grid(row=3, column=1, sticky="ew", padx=12, pady=10)

        top.grid_columnconfigure(1, weight=1)

        mid = tk.Frame(self, bg=COLORS["card"], bd=1, relief="solid")
        mid.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        tk.Label(mid, text="Users", bg=COLORS["card"], font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(10, 6))

        cols = ("Username", "Role", "Active", "Created")
        self.tree = ttk.Treeview(mid, columns=cols, show="headings", height=10)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=170, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        btns = tk.Frame(mid, bg=COLORS["card"])
        btns.pack(fill="x", padx=12, pady=(0, 12))

        tk.Button(btns, text="Disable", command=lambda: self._toggle(False)).pack(side="left")
        tk.Button(btns, text="Enable", command=lambda: self._toggle(True)).pack(side="left", padx=(8, 0))
        tk.Button(btns, text="Reset Password", command=self._reset_password).pack(side="left", padx=(8, 0))
        tk.Button(btns, text="Make Admin", command=lambda: self._set_role("admin")).pack(side="right")
        tk.Button(btns, text="Make Tech", command=lambda: self._set_role("tech")).pack(side="right", padx=(0, 8))
        tk.Button(btns, text="Rename", command=self._rename_user).pack(side="right", padx=(0, 8))

    def _selected_username(self):
        sel = self.tree.selection()
        if not sel:
            return ""
        vals = self.tree.item(sel[0], "values")
        return vals[0] if vals else ""

    def _refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for u, r, a, c in self.db.list_users():
            self.tree.insert("", "end", values=(u, r, "Yes" if a else "No", c))

    def _create_user(self, role: str):
        u = self.var_user.get().strip()
        p = self.var_pass.get()
        if not u or not p:
            messagebox.showwarning("Missing", "Username and password required")
            return
        try:
            self.db.create_user(u, p, role=role)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        self.var_user.set("")
        self.var_pass.set("")
        self._refresh()

    def _toggle(self, active: bool):
        u = self._selected_username()
        if not u:
            return
        self.db.set_user_active(u, active)
        self._refresh()

    def _reset_password(self):
        u = self._selected_username()
        if not u:
            return
        newp = simpledialog.askstring("Reset Password", f"New password for {u}:", show="*")
        if not newp:
            return
        self.db.reset_password(u, newp)
        messagebox.showinfo("OK", "Password updated")
        self._refresh()

    def _set_role(self, role: str):
        u = self._selected_username()
        if not u:
            return
        try:
            self.db.set_role(u, role)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        self._refresh()

    def _rename_user(self):
        old = self._selected_username()
        if not old:
            return
        new = simpledialog.askstring("Rename Username", f"New username for {old}:")
        if not new:
            return
        try:
            self.db.rename_user(old, new)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        messagebox.showinfo("OK", "Username updated")
        self._refresh()


class AdminPanelFrame(tk.Frame):
    def __init__(self, master, db: Database):
        super().__init__(master, bg=COLORS["bg"])
        self.db = db
        self.var_user = tk.StringVar()
        self.var_pass = tk.StringVar()
        self._build()
        self._refresh()

    def _build(self):
        wrapper = tk.Frame(self, bg=COLORS["bg"])
        wrapper.pack(fill="both", expand=True, padx=14, pady=14)

        top = tk.Frame(wrapper, bg=COLORS["card"], bd=1, relief="solid")
        top.pack(fill="x")
        tk.Label(top, text="Admin Console", bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=14, pady=(14, 4))
        tk.Label(top, text="Create and manage technician/admin accounts with a cleaner workflow.", bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=14, pady=(0, 12))

        form_area = tk.Frame(wrapper, bg=COLORS["card"], bd=1, relief="solid")
        form_area.pack(fill="x", pady=(12, 0))
        form_area.grid_columnconfigure(1, weight=1)

        tk.Label(form_area, text="Username", bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6))
        tk.Entry(form_area, textvariable=self.var_user, font=("Segoe UI", 10)).grid(row=0, column=1, sticky="ew", padx=14, pady=(12, 6))

        tk.Label(form_area, text="Password", bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", padx=14, pady=6)
        tk.Entry(form_area, textvariable=self.var_pass, show="*", font=("Segoe UI", 10)).grid(row=1, column=1, sticky="ew", padx=14, pady=6)

        tk.Label(form_area, text="Role", bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).grid(row=2, column=0, sticky="w", padx=14, pady=6)
        self.var_role = tk.StringVar(value="tech")
        role_frame = tk.Frame(form_area, bg=COLORS["card"])
        role_frame.grid(row=2, column=1, sticky="w", padx=14, pady=6)
        tk.Radiobutton(role_frame, text="Technician", variable=self.var_role, value="tech", bg=COLORS["card"], fg=COLORS["muted"], selectcolor=COLORS["card"], font=("Segoe UI", 10)).pack(side="left")
        tk.Radiobutton(role_frame, text="Admin", variable=self.var_role, value="admin", bg=COLORS["card"], fg=COLORS["muted"], selectcolor=COLORS["card"], font=("Segoe UI", 10)).pack(side="left", padx=(12, 0))

        action_frame = tk.Frame(form_area, bg=COLORS["card"])
        action_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=14, pady=(10, 14))
        tk.Button(action_frame, text="Create User", bg=COLORS["accent"], fg="white", relief="flat", command=self._create_user, width=18).pack(side="left")
        self.admin_message = tk.Label(action_frame, text="", bg=COLORS["card"], fg=COLORS["pass"], font=("Segoe UI", 10), anchor="w")
        self.admin_message.pack(side="left", padx=(16, 0), fill="x", expand=True)

        list_area = tk.Frame(wrapper, bg=COLORS["card"], bd=1, relief="solid")
        list_area.pack(fill="both", expand=True, pady=(14, 0))
        tk.Label(list_area, text="User Accounts", bg=COLORS["card"], font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=14, pady=(12, 6))

        cols = ("Username", "Role", "Active", "Created")
        self.tree = ttk.Treeview(list_area, columns=cols, show="headings", height=12)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=155, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=14, pady=(0, 10))

        btns = tk.Frame(list_area, bg=COLORS["card"])
        btns.pack(fill="x", padx=14, pady=(0, 14))
        tk.Button(btns, text="Enable", command=lambda: self._toggle(True), width=14).pack(side="left")
        tk.Button(btns, text="Disable", command=lambda: self._toggle(False), width=14).pack(side="left", padx=(8, 0))
        tk.Button(btns, text="Reset Password", command=self._reset_password, width=14).pack(side="left", padx=(8, 0))
        tk.Button(btns, text="Rename", command=self._rename_user, width=14).pack(side="left", padx=(8, 0))
        tk.Button(btns, text="Make Tech", command=lambda: self._set_role("tech"), width=14).pack(side="right")
        tk.Button(btns, text="Make Admin", command=lambda: self._set_role("admin"), width=14).pack(side="right", padx=(0, 8))

    def _selected_username(self):
        sel = self.tree.selection()
        if not sel:
            return ""
        vals = self.tree.item(sel[0], "values")
        return vals[0] if vals else ""

    def _refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for u, r, a, c in self.db.list_users():
            self.tree.insert("", "end", values=(u, r, "Yes" if a else "No", c))

    def _create_user(self, role: str = None):
        u = self.var_user.get().strip()
        p = self.var_pass.get()
        if not u or not p:
            messagebox.showwarning("Missing", "Username and password required")
            return
        selected_role = role or getattr(self, 'var_role', tk.StringVar(value='tech')).get() or 'tech'
        try:
            self.db.create_user(u, p, role=selected_role)
            self.admin_message.config(text=f'Created {selected_role} user "{u}"')
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        self.var_user.set("")
        self.var_pass.set("")
        self._refresh()

    def _toggle(self, active: bool):
        u = self._selected_username()
        if not u:
            return
        self.db.set_user_active(u, active)
        self._refresh()

    def _reset_password(self):
        u = self._selected_username()
        if not u:
            return
        newp = simpledialog.askstring("Reset Password", f"New password for {u}:", show="*")
        if not newp:
            return
        self.db.reset_password(u, newp)
        messagebox.showinfo("OK", "Password updated")
        self._refresh()

    def _set_role(self, role: str):
        u = self._selected_username()
        if not u:
            return
        try:
            self.db.set_role(u, role)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        self._refresh()

    def _rename_user(self):
        u = self._selected_username()
        if not u:
            return
        new = simpledialog.askstring("Rename Username", f"New username for {u}:")
        if not new:
            return
        try:
            self.db.rename_user(u, new)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        messagebox.showinfo("OK", "Username updated")
        self._refresh()

# =========================
# Main App
# =========================

class DNTesterStationApp(tk.Tk):
    def __init__(self):
        super().__init__()
        # Inline login: keep window visible

        # --- Settings ---
        self.settings = load_settings()

        self.configure(bg=COLORS["bg"])
        self.title("VS Verification Station")

        # Start maximized (Windows), but NOT always-on-top
        try:
            self.state('zoomed')
        except Exception:
            self.geometry("1180x720")

        # Prevent layout collapse (login should never overlap)
        try:
            self.minsize(980, 620)
        except Exception:
            pass

        # UI vars must exist before menus use them
        self.always_on_top = tk.BooleanVar(value=bool(self.settings.get('always_on_top', False)))
        self.auto_export_shared = tk.BooleanVar(value=bool(self.settings.get('auto_export_shared', True)))
        # 'flat' = one shared folder only; 'date' = YYYY-MM-DD subfolders
        self.shared_folder_mode = tk.StringVar(value=str(self.settings.get('shared_folder_mode', 'flat')))

        try:
            self.attributes("-topmost", bool(self.always_on_top.get()))
        except Exception:
            pass

        self._build_menubar()
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        self.db = Database(DB_PATH)

        if not self.db.has_any_admin():
            dlg = CreateAdminDialog(self, self.db)
            self.wait_window(dlg)
            if not dlg.created:
                self.destroy(); return
        # Inline login (no popup)
        self.tester = None
        self.role = None

        self._show_login_screen()
        return  # wait for user login

        self.current_trc_path = ""
        self.fr_win = None
        self.current_vs = None
        self.start_time = ""
        self.current_trc_time = ""
        self._trc_cache = {"path": None, "mtime": None, "vs": {}}

        self.material_var = tk.StringVar()
        self.serial_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")


        # TRC selection
        self.trc_mode = tk.StringVar(value="auto")  # auto | manual
        self.trc_path_var = tk.StringVar(value="")
        self._build_ui()
        self._load_table()
        # Hotkey: Logout
        self.bind_all("<Control-l>", lambda e: self.logout())
        self._set_result_buttons(False)



    def _clear_root_widgets(self):
        for w in self.winfo_children():
            try:
                w.destroy()
            except Exception:
                pass

    def _show_login_screen(self):
        self._clear_root_widgets()
        try:
            self.deiconify()
        except Exception:
            pass
        try:
            self.attributes('-topmost', bool(self.always_on_top.get()))
        except Exception:
            pass

        # Full-page login layout (responsive – prevents left panel overlap on small windows)
        root = tk.Frame(self, bg=COLORS['bg'])
        root.pack(fill='both', expand=True)

        # Use grid so panels never overlap; left stays fixed width, right takes remaining space
        root.grid_rowconfigure(0, weight=1)
        root.grid_columnconfigure(0, weight=0)
        root.grid_columnconfigure(1, weight=1)

        left = tk.Frame(root, bg=COLORS['header'], width=420)
        left.grid(row=0, column=0, sticky='nsew')
        try:
            left.grid_propagate(False)  # keep fixed width
        except Exception:
            pass

        right = tk.Frame(root, bg=COLORS['bg'])
        right.grid(row=0, column=1, sticky='nsew')

        tk.Label(left, text='DN Tester Station', bg=COLORS['header'], fg=COLORS['header_text'],
                 font=('Segoe UI', 26, 'bold')).pack(padx=30, pady=(70, 10), anchor='w')
        tk.Label(left, text='Diebold Nixdorf', bg=COLORS['header'], fg=COLORS['header_text'],
                 font=('Segoe UI', 14)).pack(padx=30, anchor='w')
        tk.Label(left, text='Login to continue', bg=COLORS['header'], fg=COLORS['header_text'],
                 font=('Segoe UI', 12)).pack(padx=30, pady=(30, 0), anchor='w')

        # Card uses place, but dynamically clamps to the right panel size
        card = tk.Frame(right, bg=COLORS['card'], bd=1, relief='solid')
        card.place(relx=0.5, rely=0.5, anchor='center')

        def _resize_login_card(event=None):
            try:
                rw = right.winfo_width()
                rh = right.winfo_height()
                # Clamp card size so it always fits inside right panel
                w = max(340, min(540, rw - 48))
                h = max(360, min(460, rh - 48))
                card.place_configure(width=w, height=h)
            except Exception:
                pass

        try:
            right.bind('<Configure>', _resize_login_card)
            self.after(50, _resize_login_card)
        except Exception:
            pass

        tk.Label(card, text='Login', bg=COLORS['card'], font=('Segoe UI', 18, 'bold')).pack(pady=(24, 14))

        self.login_user = tk.StringVar()
        self.login_pw = tk.StringVar()
        self.login_guest = tk.StringVar()
        self.login_msg = tk.StringVar(value='')

        frm = tk.Frame(card, bg=COLORS['card'])
        frm.pack(fill='x', padx=26)

        tk.Label(frm, text='Username', bg=COLORS['card'], fg=COLORS['muted']).grid(row=0, column=0, sticky='w')
        tk.Entry(frm, textvariable=self.login_user, font=('Segoe UI', 11)).grid(row=1, column=0, sticky='ew', pady=(3, 12))

        tk.Label(frm, text='Password', bg=COLORS['card'], fg=COLORS['muted']).grid(row=2, column=0, sticky='w')
        tk.Entry(frm, textvariable=self.login_pw, show='*', font=('Segoe UI', 11)).grid(row=3, column=0, sticky='ew', pady=(3, 14))

        frm.grid_columnconfigure(0, weight=1)

        btns = tk.Frame(card, bg=COLORS['card'])
        btns.pack(fill='x', padx=26)

        tk.Button(btns, text='Login', bg=COLORS['accent'], fg='white', relief='flat',
                  font=('Segoe UI', 11, 'bold'), command=self._do_login).pack(side='left', expand=True, fill='x', padx=(0, 6))
        tk.Button(btns, text='Guest', bg=COLORS['gray'], fg='white', relief='flat',
                  font=('Segoe UI', 11, 'bold'), command=self._do_guest).pack(side='left', expand=True, fill='x', padx=(6, 0))

        sep = tk.Frame(card, bg=COLORS['border'], height=1)
        sep.pack(fill='x', padx=26, pady=14)

        tk.Label(card, text='Guest name (required if Guest)', bg=COLORS['card'], fg=COLORS['muted']).pack(anchor='w', padx=26)
        tk.Entry(card, textvariable=self.login_guest, font=('Segoe UI', 11)).pack(fill='x', padx=26, pady=(4, 10))

        tk.Label(card, textvariable=self.login_msg, bg=COLORS['card'], fg=COLORS['danger'],
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w', padx=26)

        try:
            self.unbind('<Return>')
        except Exception:
            pass
        self.bind('<Return>', lambda e: self._do_login())

    def _init_runtime_state(self):
        self.current_trc_path = ''
        self.fr_win = None
        self.current_vs = None
        self.start_time = ''
        self.current_trc_time = ''

        self.material_var = tk.StringVar()
        self.serial_var = tk.StringVar()
        self.status_var = tk.StringVar(value='Ready')

        self.trc_mode = tk.StringVar(value='auto')
        self.trc_path_var = tk.StringVar(value='')

        # --- Dashboard / station info ---
        self.machine_name = platform.node() or 'PC'
        self.dash_today_total = tk.StringVar(value='0')
        self.dash_today_pass = tk.StringVar(value='0')
        self.dash_today_rework = tk.StringVar(value='0')
        self.dash_today_doa = tk.StringVar(value='0')
        self.dash_pass_rate = tk.StringVar(value='0%')
        self.dash_shared_status = tk.StringVar(value='Shared: not configured')
        self.dash_master_path = tk.StringVar(value='Master: not configured')
        self.dash_last_append = tk.StringVar(value='Last append: -')
        self.dash_last_error = tk.StringVar(value='')
        self.vs_status_var = tk.StringVar(value='')

    def _do_login(self):
        try:
            self.unbind('<Return>')
        except Exception:
            pass
        u = (self.login_user.get() or '').strip()
        p = self.login_pw.get() or ''
        if not u or not p:
            self.login_msg.set('Enter username and password')
            return
        role = self.db.authenticate(u, p)
        if not role:
            self.login_msg.set('Invalid credentials')
            return
        self.tester = u
        self.role = role
        self._clear_root_widgets()
        self._init_runtime_state()
        self._build_ui()
        self._load_table()
        self._set_result_buttons(False)
        self.bind_all('<Control-l>', lambda e: self.logout())
        self.deiconify(); self.lift(); self.focus_force()

    def _do_guest(self):
        try:
            self.unbind('<Return>')
        except Exception:
            pass
        n = (self.login_guest.get() or '').strip()
        if not n:
            self.login_msg.set('Guest name required')
            return
        self.tester = f'Guest - {n}'
        self.role = 'guest'
        self._clear_root_widgets()
        self._init_runtime_state()
        self._build_ui()
        self._load_table()
        self._set_result_buttons(False)
        self.bind_all('<Control-l>', lambda e: self.logout())
        self.deiconify(); self.lift(); self.focus_force()

    def _on_table_right_click(self, event):
        item = self.table.identify_row(event.y)
        if item:
            self.table.selection_set(item)
            try:
                self.menu.post(event.x_root, event.y_root)
            except Exception:
                pass

    def update_selected_result(self, new_result: str):
        sel = self.table.selection()
        if not sel:
            return
        vals = self.table.item(sel[0], 'values')
        if not vals:
            return
        session_id = vals[0]
        try:
            self.db.update_session_result(session_id, new_result)
        except Exception as e:
            messagebox.showerror('Update failed', str(e))
            return
        self.status_var.set(f'Result updated to {new_result}')
        self._load_table()
        try:
            self.refresh_dashboard()
        except Exception:
            pass
        self._auto_export_session_if_enabled(session_id)

    def open_qualifying(self):
        try:
            ensure_process_started(QUALIFYING_EXE, QUALIFYING_PROCESS)
        except Exception:
            pass

    def focus_flight_recorder(self):
        try:
            w = find_visible_fr_window(timeout_sec=10)
            try:
                w.set_focus()
            except Exception:
                pass
        except Exception:
            pass

    def show_admin_tab(self):
        try:
            if hasattr(self, 'notebook') and hasattr(self, 'admin_tab'):
                self.notebook.select(self.admin_tab)
                self._v34_show_admin_overlay()
        except Exception:
            pass

    def _build_ui(self):
        header = tk.Frame(self, bg=COLORS["header"], height=100)
        header.pack(fill="x")
        header.pack_propagate(False)

        title_frame = tk.Frame(header, bg=COLORS["header"])
        title_frame.pack(fill="both", expand=True, padx=16, pady=(10, 8))
        tk.Label(title_frame, text="VS Verification Station", bg=COLORS["header"], fg=COLORS["header_text"], font=("Segoe UI", 18, "bold")).pack(anchor='w')
        tk.Label(title_frame, text="Detect connected VS devices, start VS1/VS2, then save PASS / REWORK / DOA.", bg=COLORS["header"], fg=COLORS["header_text"], font=("Segoe UI", 10), wraplength=580, justify='left').pack(anchor='w', pady=(4, 0))

        right_hdr = tk.Frame(header, bg=COLORS["header"])
        right_hdr.place(relx=1.0, rely=0.5, anchor='e')
        tk.Label(right_hdr, text=f"User: {self.tester} ({self.role}) | Machine: {getattr(self, 'machine_name', platform.node())}", bg=COLORS["header"], fg=COLORS["header_text"], font=("Segoe UI", 10)).pack(side="left")
        tk.Button(right_hdr, text="Logout (Ctrl+L)", command=self.logout, bg=COLORS["header"], fg=COLORS["header_text"], relief="flat", activebackground=COLORS["header"], activeforeground=COLORS["header_text"], font=("Segoe UI", 9, "bold"), padx=10, pady=4).pack(side="left", padx=(12, 0))

        body = tk.Frame(self, bg=COLORS["bg"])
        body.pack(fill="both", expand=True, padx=16, pady=16)

        left = tk.Frame(body, bg=COLORS["card"], bd=1, relief="solid")
        left.configure(width=460)
        left.pack(side="left", fill="y", padx=(0, 16))
        left.pack_propagate(False)
        l = tk.Frame(left, bg=COLORS["card"])
        l.pack(fill="both", expand=True, padx=16, pady=16)

        # --- Station input ---
        info_section = tk.LabelFrame(l, text="Station Data", bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 11, "bold"), labelanchor="n")
        info_section.pack(fill="x", pady=(0, 12))
        tk.Label(info_section, text="Material Number", bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(10, 2))
        ttk.Entry(info_section, textvariable=self.material_var, width=42).pack(anchor="w", padx=10, pady=(0, 10))
        tk.Label(info_section, text="Serial Number", bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(0, 2))
        ttk.Entry(info_section, textvariable=self.serial_var, width=42).pack(anchor="w", padx=10, pady=(0, 10))

        # --- Current session ---
        session_section = tk.LabelFrame(l, text="Current Session", bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 11, "bold"), labelanchor="n")
        session_section.pack(fill="x", pady=(0, 12))
        tk.Label(session_section, text='Selected VS', bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(10, 2))
        tk.Label(session_section, textvariable=self.current_test_var, bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=10)
        tk.Label(session_section, text='Source', bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(8, 2))
        tk.Label(session_section, textvariable=self.current_source_var, bg=COLORS["card"], fg=COLORS["gray"], font=("Segoe UI", 10), wraplength=400, justify='left').pack(anchor="w", padx=10)
        tk.Label(session_section, text='Machine suggestion', bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(8, 2))
        tk.Label(session_section, textvariable=self.machine_suggest_var, bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=10)
        tk.Label(session_section, text='Confidence', bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(8, 2))
        tk.Label(session_section, textvariable=self.machine_confidence_var, bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10)
        tk.Label(session_section, text='Last save comparison', bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(8, 2))
        tk.Label(session_section, textvariable=self.last_match_status_var, bg=COLORS["card"], fg=COLORS["gray"], font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10, pady=(0, 10))

        # --- TRC selection ---
        trc_section = tk.LabelFrame(l, text="TRC Selection", bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 11, "bold"), labelanchor="n")
        trc_section.pack(fill="x", pady=(0, 12))
        mode_row = tk.Frame(trc_section, bg=COLORS["card"])
        mode_row.pack(fill="x", padx=10, pady=(10, 10))
        ttk.Radiobutton(mode_row, text="Auto (latest)", value="auto", variable=self.trc_mode).pack(side="left")
        ttk.Radiobutton(mode_row, text="Manual", value="manual", variable=self.trc_mode).pack(side="left", padx=(12, 0))
        trc_btns = tk.Frame(trc_section, bg=COLORS["card"])
        trc_btns.pack(fill="x", padx=10, pady=(0, 10))
        tk.Button(trc_btns, text="Browse TRC", command=self.browse_trc).pack(side="left")
        tk.Button(trc_btns, text="Open TRC Folder", command=self.open_trc_folder).pack(side="left", padx=(10, 0))
        tk.Label(trc_section, textvariable=self.trc_path_var, bg=COLORS["card"], fg=COLORS["gray"], font=("Segoe UI", 9), wraplength=340, justify="left").pack(anchor="w", padx=10, pady=(0, 10))

        # --- Test control ---
        test_section = tk.LabelFrame(l, text="Test Actions", bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 11, "bold"), labelanchor="n")
        test_section.pack(fill="x", pady=(0, 12))
        tk.Label(test_section, text="Start a VS session and then choose PASS/REWORK/DOA.", bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(10, 4))
        vs_bar = tk.Frame(test_section, bg=COLORS["card"])
        vs_bar.pack(fill="x", padx=10, pady=(0, 0))
        tk.Button(vs_bar, text="Start Test VS1", bg=COLORS["accent"], fg="white", relief="flat", font=("Segoe UI", 11, "bold"), padx=12, pady=8,
                  command=lambda: self.start_test("VS1")).pack(side="left", expand=True, fill="x", padx=(0, 6))
        tk.Button(vs_bar, text="Start Test VS2", bg=COLORS["accent"], fg="white", relief="flat", font=("Segoe UI", 11, "bold"), padx=12, pady=8,
                  command=lambda: self.start_test("VS2")).pack(side="left", expand=True, fill="x", padx=(6, 0))
        tk.Label(test_section, textvariable=self.vs_status_var, bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10), wraplength=340, justify="left").pack(anchor="w", padx=10, pady=(10, 4))

        result_bar = tk.Frame(test_section, bg=COLORS["card"])
        result_bar.pack(fill="x", padx=10, pady=(2, 10))
        self.btn_pass = tk.Button(result_bar, text="PASS", bg=COLORS["pass"], fg="white", relief="flat", font=("Segoe UI", 11, "bold"), padx=12, pady=8, state="disabled",
                                  command=lambda: self.finish_test("PASS"))
        self.btn_pass.pack(side="left", expand=True, fill="x", padx=(0, 6))
        self.btn_rework = tk.Button(result_bar, text="REWORK", bg=COLORS["rework"], fg="white", relief="flat", font=("Segoe UI", 11, "bold"), padx=12, pady=8, state="disabled",
                                    command=lambda: self.finish_test("REWORK"))
        self.btn_rework.pack(side="left", expand=True, fill="x", padx=6)
        self.btn_doa = tk.Button(result_bar, text="DOA", bg=COLORS["doa"], fg="white", relief="flat", font=("Segoe UI", 11, "bold"), padx=12, pady=8, state="disabled",
                                 command=lambda: self.finish_test("DOA"))
        self.btn_doa.pack(side="left", expand=True, fill="x", padx=(6, 0))
        tk.Label(test_section, textvariable=self.status_var, bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=(0, 10))

        # --- Export and admin ---
        exp_section = tk.LabelFrame(l, text="Shared Master", bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 11, "bold"), labelanchor="n")
        exp_section.pack(fill="x", pady=(0, 12))
        ttk.Button(exp_section, text="Set Shared Folder", command=self.configure_shared_folder).pack(side="left", padx=10, pady=10)
        ttk.Button(exp_section, text="Set Master File", command=self.configure_shared_master_file).pack(side="left", padx=(8, 0), pady=10)
        ttk.Button(exp_section, text="Append Visible Results → Master", command=self.export_append_table_to_master).pack(fill="x", padx=10, pady=(0, 8))
        ttk.Button(exp_section, text="Open Master", command=self.open_shared_master).pack(fill="x", padx=10, pady=(0, 10))

        if self.role == "admin":
            tk.Button(exp_section, text="Admin Panel", command=self.open_admin_panel, bg=COLORS["accent"], fg="white", relief="flat", font=("Segoe UI", 10, "bold"), padx=12, pady=8).pack(anchor="w", padx=10, pady=(0, 10))

        right = tk.Frame(body, bg=COLORS["card"], bd=1, relief="solid")
        right.pack(side="right", fill="both", expand=True, padx=(16, 0))

        self.notebook = ttk.Notebook(right)
        self.notebook.pack(fill="both", expand=True)

        # --- Dashboard tab ---
        dashboard_tab = tk.Frame(self.notebook, bg=COLORS["card"])
        self.notebook.add(dashboard_tab, text="Dashboard")
        self._build_dashboard_tab(dashboard_tab)

        # --- Results tab ---
        results_tab = tk.Frame(self.notebook, bg=COLORS["card"])
        self.notebook.add(results_tab, text="Recent Results")

        # --- Help tab ---
        help_tab = tk.Frame(self.notebook, bg=COLORS["card"])
        self.notebook.add(help_tab, text="Help / How-To")

        ht = tk.Frame(help_tab, bg=COLORS["card"])
        ht.pack(fill="both", expand=True, padx=16, pady=16)

        tk.Label(ht, text="Help / How-To", bg=COLORS["card"], fg=COLORS["header"], font=("Segoe UI", 16, "bold")).pack(anchor="w")
        tk.Label(ht, text="JDD Log Folder:\n" + str(JDD_LOG_DIR), bg=COLORS["card"], fg=COLORS["muted"], justify="left").pack(anchor="w", pady=(6, 10))

        help_msg = (
            "• The app selects the newest JDD log that contains VS identity snapshot (tcnFru=01750200435).\n"
            "• Make sure the chosen log has >=2 VS entries (VS1 + VS2).\n"
            "• Guest can use all features except Admin.\n"
            "• Use Help (F1) for detailed guidance."
        )
        t = tk.Text(ht, wrap="word", height=10, bd=1, relief="solid")
        t.insert("1.0", help_msg)
        t.configure(state="disabled")
        t.pack(fill="x", pady=(0, 10))

        ttk.Button(ht, text="Open JDD Folder", command=lambda: subprocess.Popen(f"explorer \"{JDD_LOG_DIR}\"" )).pack(anchor="w")


        r = tk.Frame(results_tab, bg=COLORS["card"])
        r.pack(fill="both", expand=True, padx=16, pady=16)

        cols = ("ID", "Tester", "Role", "VS", "Material", "Serial", "Result", "Start", "End", "TRC", "TRC_Time")
        self.table = ttk.Treeview(r, columns=cols, show="headings", height=12)

        for col in cols:
            self.table.heading(col, text=(col if col != "ID" else ""))
            width = 120
            if col == "ID":
                width = 0
            elif col == "TRC":
                width = 280
            elif col in ("Material", "Serial"):
                width = 140
            elif col in ("Start", "End"):
                width = 160
            elif col == "TRC_Time":
                width = 110
            elif col == "Role":
                width = 90
            self.table.column(col, width=width, anchor="center")

        self.table.pack(fill="both", expand=True, pady=(0, 10))

        # Right-click menu: change result
        self.menu = tk.Menu(self, tearoff=0)
        self.menu.add_command(label="Set PASS", command=lambda: self.update_selected_result("PASS"))
        self.menu.add_command(label="Set REWORK", command=lambda: self.update_selected_result("REWORK"))
        self.menu.add_command(label="Set DOA", command=lambda: self.update_selected_result("DOA"))
        self.table.bind("<Button-3>", self._on_table_right_click)

        act = tk.Frame(r, bg=COLORS["card"])
        act.pack(fill="x")
        tk.Label(act, text="Change selected result:", bg=COLORS["card"], fg=COLORS["muted"]).pack(side="left")
        self.result_choice = ttk.Combobox(act, values=["PASS", "REWORK", "DOA"], width=10, state="readonly")
        self.result_choice.set("PASS")
        self.result_choice.pack(side="left", padx=(8, 8))
        tk.Button(act, text="Apply", command=lambda: self.update_selected_result(self.result_choice.get())).pack(side="left")


        # --- Export buttons (shared master) ---
        ttk.Button(act, text="Append → Master", command=self.export_append_table_to_master).pack(side="right")
        ttk.Button(act, text="Open Master", command=self.open_shared_master).pack(side="right", padx=(8, 8))
        # --- Admin tab ---
        self.admin_tab = tk.Frame(self.notebook, bg=COLORS["bg"])
        self.notebook.add(self.admin_tab, text="Admin")
        AdminPanelFrame(self.admin_tab, self.db).pack(fill="both", expand=True)


    def open_admin_panel(self):
        try:
            self.show_admin_tab()
        except Exception:
            pass

    def _set_result_buttons(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        self.btn_pass.configure(state=state)
        self.btn_rework.configure(state=state)
        self.btn_doa.configure(state=state)

    def _clear_fields(self):
        self.material_var.set("")
        self.serial_var.set("")
        self.current_vs = None
        self.start_time = ""
        self.current_trc_time = ""
        self._set_result_buttons(False)

    def _load_table(self):
        for item in self.table.get_children():
            self.table.delete(item)
        for row in self.db.fetch_recent(limit=200):
            session_id, tester, role, vs, material, serial, result, start, end, trc, trc_time = row
            self.table.insert("", "end", values=(session_id, tester, role, vs, material, serial, result, start, end, trc, trc_time))


    def _get_vs_from_trc_cache(self, trc_path: str):
        """Extract VS mapping from TRC using cache to avoid repeated decompress."""
        try:
            p = Path(trc_path)
            mtime = p.stat().st_mtime
        except Exception:
            mtime = None
        try:
            cache = self._trc_cache
        except Exception:
            cache = {"path": None, "mtime": None, "vs": {}}
            self._trc_cache = cache
        if cache.get('path') == trc_path and cache.get('mtime') == mtime and cache.get('vs'):
            return cache['vs']
        vsmap = extract_vs_from_trc(trc_path)
        self._trc_cache = {"path": trc_path, "mtime": mtime, "vs": vsmap}
        return vsmap

    def start_test(self, vs: str):
        try:
            self._clear_fields()
            self.current_vs = vs
            self.start_time = now_ts()
            self.vs_status_var.set(f'Target: {vs}')
            self.status_var.set(f'Starting {vs}...')

            # LOG-ONLY mode: extract VS1/VS2 from JDD logs (no Flight Recorder UI)
            if USE_JDD_LOG_EXTRACTION:
                self._start_vs_from_logs(vs)
                return

            # Improve UIA reliability: optionally minimize this app while reading from FR
            was_topmost = False
            try:
                was_topmost = bool(self.attributes("-topmost"))
            except Exception:
                was_topmost = False
            if AUTO_MINIMIZE_DURING_READ:
                try:
                    self.attributes("-topmost", False)
                except Exception:
                    pass

            if AUTO_OPEN_QUALIFYING and QUALIFYING_EXE:
                ensure_process_started(QUALIFYING_EXE, QUALIFYING_PROCESS)
            # TRC path selection
            if self.trc_mode.get() == "manual" and self.trc_path_var.get().strip():
                self.current_trc_path = self.trc_path_var.get().strip()

                # TRC extraction (fast path): UI-free, deterministic read
                if USE_TRC_EXTRACTION:
                    try:
                        vsmap = self._get_vs_from_trc_cache(self.current_trc_path)
                        if vs in vsmap:
                            mat, sn = vsmap[vs]
                            self.current_trc_time = now_ts()
                            self.material_var.set(mat)
                            self.serial_var.set(sn)
                            self.status_var.set(f"Extracted from TRC ({vs}): {mat} / {sn}")
                            self._set_result_buttons(True)
                            return
                    except Exception:
                        pass
            if not self.current_trc_path:
                self.current_trc_path = pick_latest_trc(TRC_DIR)

                # TRC extraction (fast path): UI-free, deterministic read
                if USE_TRC_EXTRACTION:
                    try:
                        vsmap = self._get_vs_from_trc_cache(self.current_trc_path)
                        if vs in vsmap:
                            mat, sn = vsmap[vs]
                            self.current_trc_time = now_ts()
                            self.material_var.set(mat)
                            self.serial_var.set(sn)
                            self.status_var.set(f"Extracted from TRC ({vs}): {mat} / {sn}")
                            self._set_result_buttons(True)
                            return
                    except Exception:
                        pass
                open_trc_via_windows(self.current_trc_path)
                time.sleep(WAIT_AFTER_OPEN_TRC_SEC)
                self.fr_win = find_visible_fr_window()
            else:
                if self.fr_win is None:
                    self.fr_win = find_visible_fr_window()
                try:
                    self.fr_win.set_focus()
                except Exception:
                    pass
                click_reload_if_present(self.fr_win)

            self.status_var.set(f"Using TRC: {Path(self.current_trc_path).name} | {vs}")
            self.update_idletasks()

            sysinfo = find_tree_like_item(self.fr_win, SYSINFO_TEXT) or find_control_by_text(self.fr_win, SYSINFO_TEXT)
            if not sysinfo:
                sysinfo = find_control_contains(self.fr_win, "system", "info") or find_control_contains(self.fr_win, "system")
            if not sysinfo:
                raise RuntimeError("System Info control not found")

            safe_invoke(sysinfo)
            time.sleep(0.8)

            self.current_trc_time = select_latest_timestamp(self.fr_win)
            time.sleep(0.4)

            # Select VS and read values (retry once if unstable on first run)
            mat = sn = ""
            for attempt in range(2):
                try:
                    try:
                        self.fr_win.set_focus()
                    except Exception:
                        pass
                    time.sleep(0.4)

                    select_vs_board(self.fr_win, vs)
                    time.sleep(0.8)

                    # FAST (verified): copy clipboard and confirm objectId matches selected VS
                    expected_obj = VS_BOARD_TEXT.get(vs, '')
                    for _try in range(3):
                        if self._copy_details_to_clipboard():
                            time.sleep(0.12)
                            obj, mat, sn = self._read_fru_from_clipboard()
                            if expected_obj and expected_obj in (obj or '') and mat and sn:
                                self.material_var.set(mat)
                                self.serial_var.set(sn)
                                self.status_var.set(f"Loaded {vs}: {mat} / {sn} @ {self.current_trc_time}")
                                self._set_result_buttons(True)
                                return
                        # selection not updated yet -> reselect and retry
                        try:
                            select_vs_board(self.fr_win, vs)
                        except Exception:
                            pass
                        time.sleep(0.25)

                    # Fallback: old UI scan
                    mat_raw = read_key_value(self.fr_win, KEY_MAT)
                    sn_raw = read_key_value(self.fr_win, KEY_SN)
                    mat = clean_material(mat_raw)
                    sn = (sn_raw.strip().split()[0] if sn_raw else "")
                    if mat and sn:
                        break
                except Exception:
                    pass
                # extra refresh and retry
                try:
                    click_reload_if_present(self.fr_win)
                except Exception:
                    pass
                time.sleep(0.8)


            if not mat or not sn:
                raise RuntimeError("Could not read tcnFru/snoFru")

            self.material_var.set(mat)
            self.serial_var.set(sn)
            self.status_var.set(f"Loaded {vs}: {mat} / {sn} @ {self.current_trc_time}")
            self._set_result_buttons(True)

            try:
                self.deiconify()
                self.lift(); self.focus_force()
                self.attributes("-topmost", was_topmost if was_topmost else True)
            except Exception:
                pass

        except Exception as e:
            try:
                self.deiconify()
                self.lift(); self.focus_force()
                self.attributes("-topmost", was_topmost if was_topmost else True)
            except Exception:
                pass
            messagebox.showerror("Start Test failed", str(e))
            self.status_var.set(f"Error: {e}")


    # -------- Live JDD log wait (VS1/VS2, LOG-ONLY) --------
    def _poll_vs_from_logs(self, vs: str, baseline_mtime: float, started_at: float,
                           last_path: str = '', last_size: int = -1, stable: int = 0):
        if (time.time() - started_at) > JDD_LOG_TIMEOUT_SEC:
            self.status_var.set('Timeout: VS data not found in JDD logs (check JDD_LOG_DIR and VS*_SNO_ELEC_CANDIDATES)')
            return

        log_path = _jdd_find_new_log(JDD_LOG_DIR, baseline_mtime) if JDD_REQUIRE_NEW_LOG else ''
        if not log_path:
            # fallback to latest existing log (offline / or if live keeps writing same file)
            log_path = _jdd_pick_best_log_for_vs_material(JDD_LOG_DIR, baseline_mtime if JDD_REQUIRE_NEW_LOG else 0.0, target_vs=vs)

        if not log_path:
            self.vs_status_var.set(f'Waiting for {vs} data...')
            self.after(JDD_LOG_POLL_MS, lambda: self._poll_vs_from_logs(vs, baseline_mtime, started_at,
                                                                        last_path, last_size, stable))
            return

        try:
            size = Path(log_path).stat().st_size
        except Exception:
            size = -1

        if log_path == last_path and size == last_size:
            stable += 1
        else:
            stable = 0

        if stable < JDD_LOG_STABLE_PASSES:
            self.vs_status_var.set(f'Found log {Path(log_path).name}. Waiting for stable {vs} data...')
            self.after(JDD_LOG_POLL_MS, lambda: self._poll_vs_from_logs(vs, baseline_mtime, started_at,
                                                                        log_path, size, stable))
            return

        try:
            vsmap = extract_vs_from_jdd_log_file(log_path)
        except Exception:
            vsmap = {}

        if vs in vsmap:
            mat, sn = vsmap[vs]
            self.material_var.set(mat)
            self.serial_var.set(sn)
            # Reuse TRC fields as evidence pointer (keeps DB schema unchanged)
            self.current_trc_path = log_path
            self.current_trc_time = now_ts()
            self.vs_status_var.set(f'Loaded {vs} from {Path(log_path).name}')
            self.status_var.set(f"(Log) Loaded {vs}: {mat} / {sn} (from {Path(log_path).name})")
            self._set_result_buttons(True)
            return

        if vsmap:
            self.vs_status_var.set(f'Log {Path(log_path).name} contains: {", ".join(sorted(vsmap.keys()))}')
        else:
            self.vs_status_var.set(f'Log {Path(log_path).name} does not yet contain {vs}')

        self.after(JDD_LOG_POLL_MS, lambda: self._poll_vs_from_logs(vs, baseline_mtime, started_at,
                                                                    log_path, size, stable))

    def _start_vs_from_logs(self, vs: str):
        baseline = _jdd_latest_log_mtime(JDD_LOG_DIR) if JDD_REQUIRE_NEW_LOG else 0.0
        self.status_var.set(f"Waiting for VS data (smart newest VS-log) ({vs})...")
        self.update_idletasks()
        self._poll_vs_from_logs(vs, baseline, time.time())
        return


    def finish_test(self, result: str):
        if not self.material_var.get().strip() or not self.serial_var.get().strip():
            messagebox.showerror("Missing data", "Material and Serial required")
            return
        if not self.current_trc_path:
            messagebox.showerror("Missing TRC", "TRC path missing")
            return

        end_time = now_ts()
        sid = self.db.insert_session(
            tester=self.tester,
            role=self.role,
            vs=self.current_vs or "",
            material=self.material_var.get().strip(),
            serial=self.serial_var.get().strip(),
            result=result,
            start_time=self.start_time or end_time,
            end_time=end_time,
            trc_path=self.current_trc_path,
            trc_time=self.current_trc_time,
        )

        try:
            self.db.save_trc_blob(sid, self.current_trc_path)
        except Exception as exc:
            messagebox.showwarning("TRC save", f"Saved result, but failed to store TRC in DB:\n{exc}")

        self.status_var.set(f"Saved: {result} ({self.current_vs})")
        self._load_table()
        try:
            self.refresh_dashboard()
        except Exception:
            pass

        # Auto-export (shared master) after saving a result
        self._auto_export_session_if_enabled(sid)

        self.current_trc_path = ""
        self.fr_win = None
        self._clear_fields()

    def _copy_details_to_clipboard(self) -> bool:
        """Right-click in Flight Recorder *details pane* and choose Copy to clipboard.",
        Returns True if we likely triggered the copy."""
        if self.fr_win is None or Desktop is None:
            return False
        # clear clipboard to detect fresh copy
        try:
            self.clipboard_clear()
        except Exception:
            pass
        
        try:
            try:
                self.fr_win.set_focus()
            except Exception:
                pass
            time.sleep(0.12)
        
            # Click inside details pane (right side) to ensure correct context menu
            try:
                r = self.fr_win.rectangle()
                # 75% width, 45% height tends to be inside the value/details grid, not the left tree
                x_rel = int(r.width() * 0.75)
                y_rel = int(r.height() * 0.45)
                # focus details with left click
                try:
                    self.fr_win.click_input(button="left", coords=(x_rel, y_rel))
                except Exception:
                    pass
                time.sleep(0.06)
                # open context menu
                self.fr_win.click_input(button="right", coords=(x_rel, y_rel))
            except Exception:
                # fallback: just right click
                self.fr_win.click_input(button="right")
        
            time.sleep(0.15)
        
            desk = Desktop(backend="uia")
            # Try to locate a Menu and click the item
            try:
                menu = desk.window(control_type="Menu")
                menu.wait("visible", timeout=1.0)
                item = menu.child_window(title_re=r"(?i)copy\s*to\s*clipboard|copy.*clipboard", control_type="MenuItem")
                try:
                    item.invoke()
                except Exception:
                    item.click_input()
                time.sleep(0.12)
                return True
            except Exception:
                pass
        
            # Keyboard fallback: Shift+F10 then press C then Enter
            try:
                self.fr_win.type_keys("+{F10}")
                time.sleep(0.12)
                self.fr_win.type_keys("c")
                time.sleep(0.05)
                self.fr_win.type_keys("{ENTER}")
                time.sleep(0.12)
                return True
            except Exception:
                pass
        
            return False
        except Exception:
            return False
    def _read_fru_from_clipboard(self) -> tuple[str, str]:
        """Read objectId/tcnFru/snoFru from clipboard (first column)."""
        try:
            clip = self.clipboard_get()
        except Exception:
            clip = ""
        data = parse_clipboard_kv_first(clip)
        obj = data.get("objectId", "")
        mat_raw = data.get("tcnFru", "")
        sn_raw = data.get("snoFru", "")
        mat = clean_material(mat_raw)
        sn = (sn_raw.strip().split()[0] if sn_raw else "")
        return obj, mat, sn
    def browse_trc(self):
        """Pick a TRC file manually."""
        try:
            path = filedialog.askopenfilename(title="Select TRC", initialdir=TRC_DIR, filetypes=[("TRC files", "*.trc"), ("All files", "*.*")])
        except Exception:
            path = ""
        if path:
            self.trc_mode.set("manual")
            self.trc_path_var.set(path)
            # Reset cached FR window so it re-detects after opening
            self.fr_win = None
            self.current_trc_path = path

    def open_trc_folder(self):
        """Open TRC folder in Explorer."""
        try:
            folder = TRC_DIR
            p = (self.trc_path_var.get() or "").strip()
            if p:
                try:
                    folder = str(Path(p).parent)
                except Exception:
                    folder = TRC_DIR
            subprocess.Popen(f'explorer "{folder}"')
        except Exception:
            pass

    def logout(self):
        """Logout and return to inline login screen."""
        try:
            self.tester = None
            self.role = None
            self._show_login_screen()
        except Exception:
            try:
                self.destroy()
            except Exception:
                pass

    def _build_menubar(self):
        # Safety: ensure variables exist even if init order changes
        if not hasattr(self, 'settings') or not isinstance(getattr(self, 'settings', None), dict):
            self.settings = load_settings()
        if not hasattr(self, 'always_on_top'):
            self.always_on_top = tk.BooleanVar(value=bool(self.settings.get('always_on_top', False)))
        if not hasattr(self, 'auto_export_shared'):
            self.auto_export_shared = tk.BooleanVar(value=bool(self.settings.get('auto_export_shared', True)))
        if not hasattr(self, 'shared_folder_mode'):
            self.shared_folder_mode = tk.StringVar(value=str(self.settings.get('shared_folder_mode', 'flat')))

        def _cmd(name: str):
            # Never crash menu creation if a method is missing
            fn = getattr(self, name, None)
            if callable(fn):
                return fn
            return lambda: messagebox.showerror('Missing feature', f'Method not found: {name}')

        menubar = tk.Menu(self)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label='Logout (Ctrl+L)', command=_cmd('logout'))
        file_menu.add_separator()
        file_menu.add_command(label='Exit', command=self.destroy)
        menubar.add_cascade(label='File', menu=file_menu)

        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_checkbutton(label='Always on top', variable=self.always_on_top, command=_cmd('_apply_topmost'))
        menubar.add_cascade(label='View', menu=view_menu)

        export_menu = tk.Menu(menubar, tearoff=0)
        export_menu.add_command(label='Export to Excel (App folder)', command=_cmd('export_excel_app_folder'))
        export_menu.add_command(label='Export to Excel (Choose folder...)', command=_cmd('export_excel_choose_folder'))
        export_menu.add_separator()
        export_menu.add_command(label='Export to Shared Folder (new file)', command=_cmd('export_excel_shared'))
        export_menu.add_command(label='Append Results to Shared Master', command=_cmd('export_append_table_to_master'))
        menubar.add_cascade(label='Export', menu=export_menu)

        settings_menu = tk.Menu(menubar, tearoff=0)
        settings_menu.add_command(label='Shared Folder...', command=_cmd('configure_shared_folder'))
        settings_menu.add_command(label='Shared Master File...', command=_cmd('configure_shared_master_file'))
        settings_menu.add_separator()
        settings_menu.add_checkbutton(label='Auto-export to shared after result', variable=self.auto_export_shared, command=_cmd('_apply_auto_export_setting'))

        struct_menu = tk.Menu(settings_menu, tearoff=0)
        struct_menu.add_radiobutton(label='Flat folder (one master file)', variable=self.shared_folder_mode, value='flat', command=_cmd('_apply_shared_folder_mode'))
        struct_menu.add_radiobutton(label='Date folders (YYYY-MM-DD)', variable=self.shared_folder_mode, value='date', command=_cmd('_apply_shared_folder_mode'))
        settings_menu.add_cascade(label='Shared folder structure', menu=struct_menu)

        settings_menu.add_separator()
        settings_menu.add_command(label='Open Settings Folder', command=_cmd('open_settings_folder'))
        menubar.add_cascade(label='Settings', menu=settings_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label='Help (F1)', command=_cmd('show_help'))
        help_menu.add_command(label='About', command=_cmd('show_about'))
        menubar.add_cascade(label='Help', menu=help_menu)

        self.config(menu=menubar)
        try:
            self.bind('<F1>', lambda e: _cmd('show_help')())
        except Exception:
            pass

    def _apply_topmost(self):
        try:
            val = bool(self.always_on_top.get())
            self.attributes('-topmost', val)
            try:
                self.settings['always_on_top'] = val
                save_settings(self.settings)
            except Exception:
                pass
        except Exception:
            pass


        try:
            self.attributes('-topmost', bool(self.always_on_top.get()))
        except Exception:
            pass



    # =========================
    # Settings persistence actions
    # =========================
    def _apply_auto_export_setting(self):
        try:
            self.settings['auto_export_shared'] = bool(self.auto_export_shared.get())
            save_settings(self.settings)
        except Exception:
            pass

    def _apply_shared_folder_mode(self):
        try:
            self.settings['shared_folder_mode'] = str(self.shared_folder_mode.get() or 'flat')
            save_settings(self.settings)
        except Exception:
            pass

    def configure_shared_folder(self):
        """Choose shared folder where all machines write reports."""
        try:
            current = (self.settings.get('shared_report_dir') or '').strip()
        except Exception:
            current = ''

        msg = (
            'Select the shared folder where ALL machines should write the master Excel.\n\n'
            'Tip: Use a network path like: \\\\server\\share\\DNReports'
        )
        try:
            messagebox.showinfo('Shared Folder', msg)
        except Exception:
            pass

        folder = filedialog.askdirectory(title='Choose Shared Folder', initialdir=current or str(Path.home()))
        if not folder:
            return

        try:
            self.settings['shared_report_dir'] = folder
            save_settings(self.settings)
            messagebox.showinfo('Shared Folder', f'Saved shared folder:\n{folder}')
        except Exception as e:
            messagebox.showerror('Shared Folder', str(e))

    def configure_shared_master_file(self):
        """Set master Excel filename stored inside the shared folder."""
        try:
            current = (self.settings.get('shared_master_file') or 'DNTesterStation_Master.xlsx').strip()
            new = simpledialog.askstring('Master File', 'Shared master Excel filename (inside shared folder):', initialvalue=current)
            if not new:
                return
            new = sanitize_filename(new.strip())
            if not new.lower().endswith('.xlsx'):
                new += '.xlsx'
            self.settings['shared_master_file'] = new
            save_settings(self.settings)
            messagebox.showinfo('Master File', f'Saved master filename:\n{new}')
        except Exception as e:
            messagebox.showerror('Master File', str(e))

    def open_settings_folder(self):
        try:
            p = SETTINGS_PATH.parent
            p.mkdir(parents=True, exist_ok=True)
            subprocess.Popen(f'explorer "{p}"')
        except Exception:
            pass

    # =========================
    # Shared folder + Master helpers
    # =========================
    def _shared_base_dir(self) -> Path:
        p = (self.settings.get('shared_report_dir') or '').strip()
        return Path(p) if p else Path('')

    def _shared_target_dir(self) -> Path:
        base = self._shared_base_dir()
        if not base:
            return base
        mode = (self.shared_folder_mode.get() or 'flat') if hasattr(self, 'shared_folder_mode') else (self.settings.get('shared_folder_mode') or 'flat')
        if mode == 'date':
            day = datetime.now().strftime('%Y-%m-%d')
            return base / day
        return base

    def _ensure_shared_dir(self) -> Path:
        base = self._shared_base_dir()
        if not base:
            return base
        target = self._shared_target_dir()
        try:
            target.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        return target

    def _shared_master_path(self) -> Path:
        target = self._ensure_shared_dir()
        if not target:
            return Path('')
        name = (self.settings.get('shared_master_file') or 'DNTesterStation_Master.xlsx').strip() or 'DNTesterStation_Master.xlsx'
        name = sanitize_filename(name)
        if not name.lower().endswith('.xlsx'):
            name += '.xlsx'
        return target / name

    # =========================
    # Shared file lock (prevents corruption when multiple PCs write)
    # =========================
    def _acquire_shared_lock(self, lock_path: Path, timeout_sec: float = 12.0, poll_sec: float = 0.2) -> bool:
        try:
            import time
            start = time.time()
            while time.time() - start < timeout_sec:
                try:
                    fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                    try:
                        os.write(fd, f"{platform.node()} {datetime.now().isoformat()}".encode('utf-8', errors='ignore'))
                    finally:
                        os.close(fd)
                    return True
                except FileExistsError:
                    time.sleep(poll_sec)
                except Exception:
                    time.sleep(poll_sec)
            return False
        except Exception:
            return False

    def _release_shared_lock(self, lock_path: Path) -> None:
        try:
            if lock_path.exists():
                lock_path.unlink()
        except Exception:
            pass

    # =========================
    # Master workbook append (with Excel Table filters)
    # =========================
    def _append_rows_to_master(self, master_path: Path, headers, rows) -> None:
        if Workbook is None or load_workbook is None:
            raise RuntimeError('openpyxl is not installed. Install with: pip install openpyxl')

        if master_path.exists():
            wb = load_workbook(str(master_path))
            ws = wb.active
            if ws.title != 'Results':
                ws.title = 'Results'

            # Validate header compatibility (prevents corrupt/shifted columns)
            try:
                existing = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                if existing and headers and (str(existing[0] or '') != str(headers[0] or '')):
                    raise RuntimeError(
                        'Shared master header mismatch. This master file was created by an older build.\n'
                        'Fix: Settings → Shared Master File… → choose a NEW filename (e.g. DNTesterStation_Master_v2.xlsx)\n'
                        'or rename/delete the old master file, then export again.'
                    )
            except StopIteration:
                pass
            except Exception:
                raise
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Results'
            ws.append(headers)
            try:
                if Table and TableStyleInfo and get_column_letter:
                    end_col = get_column_letter(len(headers))
                    tab = Table(displayName='ResultsTable', ref=f"A1:{end_col}1")
                    style = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True, showColumnStripes=False)
                    tab.tableStyleInfo = style
                    ws.add_table(tab)
                    ws.auto_filter.ref = tab.ref
            except Exception:
                pass

        for r in rows:
            ws.append(list(r))

        try:
            if ws.tables and get_column_letter:
                t = list(ws.tables.values())[0]
                end_row = ws.max_row
                end_col = get_column_letter(len(headers))
                t.ref = f"A1:{end_col}{end_row}"
                ws.auto_filter.ref = t.ref
            elif get_column_letter:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        except Exception:
            pass

        try:
            if Font:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    if Alignment:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.freeze_panes = 'A2'
            if get_column_letter:
                for idx, header in enumerate(headers, start=1):
                    max_len = len(str(header))
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                        for cell in row:
                            if cell.value is not None:
                                max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(idx)].width = min(60, max(14, max_len + 2))
        except Exception:
            pass

        master_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(master_path))

    # =========================
    # Export helpers
    # =========================
    def _get_table_headers(self):
        if not hasattr(self, 'table'):
            return []
        cols = list(self.table['columns'])
        headers = [self.table.heading(c).get('text') or c for c in cols]
        if headers and headers[0] == '':
            headers = headers[1:]
        return headers

    def _get_table_headers_and_rows(self):
        if not hasattr(self, 'table'):
            raise RuntimeError('Results table is not available yet.')
        cols = list(self.table['columns'])
        headers = [self.table.heading(c).get('text') or c for c in cols]
        headers = [h for h in headers if h != '']
        rows = []
        for iid in self.table.get_children(''):
            vals = list(self.table.item(iid, 'values') or [])
            if len(vals) == len(cols):
                vals = vals[1:]  # drop ID
            rows.append(vals)
        return headers, rows

    def _build_export_filename(self) -> str:
        tester = sanitize_filename(getattr(self, 'tester', '') or 'Unknown')
        host = sanitize_filename(platform.node() or 'PC')
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{tester}_{host}_{ts}.xlsx"

    def _write_excel(self, dest: Path, headers, rows):
        if Workbook is None:
            raise RuntimeError('openpyxl is not installed. Install with: pip install openpyxl')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Results'
        ws.append(headers)
        for r in rows:
            ws.append(list(r))
        try:
            if Font:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    if Alignment:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.freeze_panes = 'A2'
            if get_column_letter:
                for col_idx, header in enumerate(headers, start=1):
                    max_len = len(str(header))
                    for row in rows[:1000]:
                        try:
                            max_len = max(max_len, len(str(row[col_idx-1])))
                        except Exception:
                            pass
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(14, max_len + 2))
        except Exception:
            pass
        dest.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(dest))

    def export_excel_app_folder(self):
        try:
            headers, rows = self._get_table_headers_and_rows()
            headers = ["Machine"] + headers
            rows = [[self.machine_name] + list(r) for r in rows]
            dest = Path(__file__).resolve().parent / self._build_export_filename()
            self._write_excel(dest, headers, rows)
            messagebox.showinfo('Export', f'Excel exported to:\n{dest}')
        except Exception as e:
            messagebox.showerror('Export failed', str(e))

    def export_excel_choose_folder(self):
        try:
            headers, rows = self._get_table_headers_and_rows()
            initial = self.settings.get('last_export_dir') or str(Path(__file__).resolve().parent)
            fname = self._build_export_filename()
            path = filedialog.asksaveasfilename(
                title='Export to Excel',
                defaultextension='.xlsx',
                initialdir=initial,
                initialfile=fname,
                filetypes=[('Excel files', '*.xlsx')]
            )
            if not path:
                return
            dest = Path(path)
            self._write_excel(dest, headers, rows)
            self.settings['last_export_dir'] = str(dest.parent)
            save_settings(self.settings)
            messagebox.showinfo('Export', f'Excel exported to:\n{dest}')
        except Exception as e:
            messagebox.showerror('Export failed', str(e))

    def export_excel_shared(self):
        """Export full table as a NEW file into the shared folder (not the master)."""
        try:
            target = self._ensure_shared_dir()
            if not target or not target.exists():
                raise RuntimeError('Shared folder is not configured or not accessible. Use Settings → Shared Folder.')
            headers, rows = self._get_table_headers_and_rows()
            dest = target / self._build_export_filename()
            self._write_excel(dest, headers, rows)
            messagebox.showinfo('Export', f'Excel exported to shared folder:\n{dest}')
        except Exception as e:
            messagebox.showerror('Export failed', str(e))

    def export_append_table_to_master(self):
        """Append all visible rows to the shared master workbook."""
        try:
            target = self._ensure_shared_dir()
            if not target or not target.exists():
                raise RuntimeError('Shared folder is not configured or not accessible. Use Settings → Shared Folder.')
            headers, rows = self._get_table_headers_and_rows()
            master_path = self._shared_master_path()
            if not master_path:
                raise RuntimeError('Master file path is not available.')
            lock_path = master_path.with_suffix(master_path.suffix + '.lock')
            if not self._acquire_shared_lock(lock_path, timeout_sec=20.0):
                raise RuntimeError('Shared master is busy (locked). Try again in a moment.')
            try:
                self._append_rows_to_master(master_path, headers, rows)
            finally:
                self._release_shared_lock(lock_path)
            try:
                self.dash_last_append.set(f'Last append: {now_ts()}')
                self.dash_last_error.set('')
            except Exception:
                pass
            messagebox.showinfo('Export', f'Appended {len(rows)} rows to master:\n{master_path}')
        except Exception as e:
            messagebox.showerror('Export failed', str(e))

    # =========================
    # Auto-export after result
    # =========================
    def _export_single_session_to_master(self, session_id: str):
        try:
            if not session_id:
                return
            master_path = self._shared_master_path()
            if not master_path:
                return
            target = self._ensure_shared_dir()
            if not target or not target.exists():
                return

            row = None
            try:
                row = self.db.fetch_session(session_id)
            except Exception:
                row = None
            if not row:
                return

            headers = ['Machine'] + (self._get_table_headers() or ["Tester","Role","VS","Material","Serial","Result","Start","End","TRC","TRC_Time"]) 
            data = [self.machine_name] + [row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10]]

            lock_path = master_path.with_suffix(master_path.suffix + '.lock')
            if not self._acquire_shared_lock(lock_path, timeout_sec=12.0):
                return
            try:
                self._append_rows_to_master(master_path, headers, [data])
                try:
                    self.dash_last_append.set(f'Last append: {now_ts()}')
                    self.dash_last_error.set('')
                except Exception:
                    pass
            finally:
                self._release_shared_lock(lock_path)
        except Exception:
            pass

    def _auto_export_session_if_enabled(self, session_id: str):
        try:
            if bool(self.auto_export_shared.get()):
                self._export_single_session_to_master(session_id)
        except Exception:
            pass



    # =========================
    # Dashboard
    # =========================
    def _build_dashboard_tab(self, parent):
        """High-level overview (station health, today KPIs, shared master status)."""
        wrap = tk.Frame(parent, bg=COLORS["card"])
        wrap.pack(fill="both", expand=True, padx=16, pady=16)

        tk.Label(wrap, text="Dashboard", bg=COLORS["card"], fg=COLORS["header"],
                 font=("Segoe UI", 16, "bold")).pack(anchor="w")

        # Top info row
        info = tk.Frame(wrap, bg=COLORS["card"])
        info.pack(fill="x", pady=(10, 12))

        left = tk.Frame(info, bg=COLORS["card"])
        left.pack(side="left", fill="x", expand=True)
        tk.Label(left, text=f"User: {getattr(self, 'tester', '')} ({getattr(self, 'role', '')})",
                 bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w")
        tk.Label(left, text=f"Machine: {getattr(self, 'machine_name', platform.node())}",
                 bg=COLORS["card"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", pady=(2, 0))

        right = tk.Frame(info, bg=COLORS["card"])
        right.pack(side="right")
        ttk.Button(right, text="Refresh", command=self.refresh_dashboard).pack(side="right")
        ttk.Button(right, text="Open Shared Folder", command=self.open_shared_folder).pack(side="right", padx=(0, 8))
        ttk.Button(right, text="Open Master", command=self.open_shared_master).pack(side="right", padx=(0, 8))

        # KPI cards
        cards = tk.Frame(wrap, bg=COLORS["card"])
        cards.pack(fill="x")

        def card(title, var, color):
            c = tk.Frame(cards, bg=COLORS["bg"], bd=1, relief="solid")
            c.pack(side="left", padx=(0, 10), ipadx=10, ipady=8)
            tk.Label(c, text=title, bg=COLORS["bg"], fg=COLORS["muted"], font=("Segoe UI", 9)).pack(anchor="w")
            tk.Label(c, textvariable=var, bg=COLORS["bg"], fg=color, font=("Segoe UI", 16, "bold")).pack(anchor="w")
            return c

        card("Today Total", self.dash_today_total, COLORS["header"])
        card("PASS", self.dash_today_pass, COLORS["pass"])
        card("REWORK", self.dash_today_rework, COLORS["rework"])
        card("DOA", self.dash_today_doa, COLORS["doa"])
        card("Pass Rate", self.dash_pass_rate, COLORS["accent"])

        # Shared master status
        box = tk.Frame(wrap, bg=COLORS["card"], bd=1, relief="solid")
        box.pack(fill="x", pady=(14, 0))
        box2 = tk.Frame(box, bg=COLORS["card"])
        box2.pack(fill="x", padx=12, pady=12)

        tk.Label(box2, text="Shared Master Status", bg=COLORS["card"], fg=COLORS["header"],
                 font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(box2, textvariable=self.dash_shared_status, bg=COLORS["card"], fg=COLORS["muted"], justify="left").pack(anchor="w", pady=(6, 0))
        tk.Label(box2, textvariable=self.dash_master_path, bg=COLORS["card"], fg=COLORS["muted"], justify="left").pack(anchor="w", pady=(2, 0))
        tk.Label(box2, textvariable=self.dash_last_append, bg=COLORS["card"], fg=COLORS["muted"], justify="left").pack(anchor="w", pady=(2, 0))
        if hasattr(self, 'dash_last_error'):
            tk.Label(box2, textvariable=self.dash_last_error, bg=COLORS["card"], fg=COLORS["danger"], justify="left").pack(anchor="w", pady=(6, 0))

        # initial refresh
        try:
            self.after(150, self.refresh_dashboard)
        except Exception:
            pass

    def refresh_dashboard(self):
        """Recalculate today KPIs and shared master configuration."""
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            rows = []
            try:
                rows = self.db.fetch_recent(limit=2000) if hasattr(self, 'db') else []
            except Exception:
                rows = []

            # sessions row format from fetch_recent: (session_id, tester, role, vs, material, serial, result, start_time, end_time, trc_path, trc_time)
            t_total = t_pass = t_rework = t_doa = 0
            for r in rows or []:
                start = (r[7] or '')
                if not start.startswith(today):
                    continue
                t_total += 1
                res = (r[6] or '').upper()
                if res == 'PASS':
                    t_pass += 1
                elif res == 'REWORK':
                    t_rework += 1
                elif res == 'DOA':
                    t_doa += 1

            self.dash_today_total.set(str(t_total))
            self.dash_today_pass.set(str(t_pass))
            self.dash_today_rework.set(str(t_rework))
            self.dash_today_doa.set(str(t_doa))
            rate = (t_pass / t_total * 100.0) if t_total else 0.0
            self.dash_pass_rate.set(f"{rate:.0f}%")

            shared = (self.settings.get('shared_report_dir') or '').strip()
            master = (self.settings.get('shared_master_file') or 'DNTesterStation_Master.xlsx').strip()
            if shared:
                self.dash_shared_status.set(f"Shared: {shared}")
                self.dash_master_path.set(f"Master: {master}")
            else:
                self.dash_shared_status.set("Shared: not configured")
                self.dash_master_path.set("Master: not configured")

        except Exception as e:
            try:
                self.dash_last_error.set(str(e))
            except Exception:
                pass

    # =========================
    # Shared folder helpers (UI)
    # =========================
    def open_shared_folder(self):
        try:
            shared = (self.settings.get('shared_report_dir') or '').strip()
            if not shared:
                messagebox.showerror('Shared Folder', 'Shared folder is not configured. Use Settings → Shared Folder...')
                return
            subprocess.Popen(f'explorer "{shared}"')
        except Exception:
            pass

    def open_shared_master(self):
        try:
            p = self._shared_master_path() if hasattr(self, '_shared_master_path') else None
            if not p:
                messagebox.showerror('Master', 'Shared master is not configured. Use Settings → Shared Folder/Master File...')
                return
            # Open folder and select file if exists
            if Path(p).exists():
                subprocess.Popen(f'explorer /select,"{p}"')
            else:
                subprocess.Popen(f'explorer "{Path(p).parent}"')
        except Exception:
            pass

    def show_help(self):
        win = tk.Toplevel(self)
        win.title('Help')
        win.configure(bg=COLORS['bg'])
        win.geometry('720x520')
        try:
            win.minsize(560, 420)
        except Exception:
            pass

        card = tk.Frame(win, bg=COLORS['card'], bd=1, relief='solid')
        card.pack(fill='both', expand=True, padx=14, pady=14)

        tk.Label(card, text='DN Tester Station – Help', bg=COLORS['card'], fg=COLORS['header'],
                 font=('Segoe UI', 16, 'bold')).pack(anchor='w', padx=14, pady=(14, 8))

        txt = tk.Text(card, wrap='word', height=20, bd=0)
        txt.pack(fill='both', expand=True, padx=14, pady=(0, 10))

        help_text = (
            "Quick guide:\n"
            "1) Put your JDD logs into the configured folder.\n"
            "2) The app selects the newest log containing VS identity (tcnFru=01750200435).\n"
            "3) Ensure the chosen log has >=2 VS entries (VS1 + VS2).\n\n"
            "Security:\n"
            + f"• Login lockout: {LOCKOUT_MAX_ATTEMPTS} attempts / {LOCKOUT_SECONDS}s\n\n"
            + "Configured JDD folder:\n" + str(JDD_LOG_DIR) + "\n"
        )
        txt.insert('1.0', help_text)
        txt.configure(state='disabled')

        btns = tk.Frame(card, bg=COLORS['card'])
        btns.pack(fill='x', padx=14, pady=(0, 14))
        ttk.Button(btns, text='Open JDD Folder', command=lambda: subprocess.Popen(f'explorer "{JDD_LOG_DIR}"')).pack(side='left')
        ttk.Button(btns, text='Close', command=win.destroy).pack(side='right')

        try:
            win.transient(self)
            win.grab_set()
        except Exception:
            pass

    def show_about(self):
        messagebox.showinfo('About', 'DN Tester Station\nLog-only VS extraction\nBuild: V25 Pro – Shared Master (single file)')

def main():
    # Diagnostic banner to prove this edited file is being executed.
    try:
        run_path = Path(__file__).resolve()
        print('\n' + '='*60)
        print('MODERN UI ACTIVE - DNTesterStation_MODERN_FINAL.py')
        print('Running file:', run_path)
        print('Timestamp:', datetime.now().isoformat())
        print('='*60 + '\n')
        try:
            tmp = Path('/tmp') / 'dn_modern_ui_active.txt'
            tmp.write_text(f"{run_path}\n{datetime.now().isoformat()}\n")
        except Exception:
            pass
    except Exception:
        pass
    DNTesterStationApp().mainloop()



# =========================================================
# V34 CLEAN RETHINK (based on original V26 only)
# - Keep visible PASS / REWORK / DOA buttons from original V26
# - Remove popup progress window; show test context inline in main UI
# - Detect VS1 and VS2 from recent JDD logs (not only single latest file)
# - Add machine suggestion, tester note, ML feedback, and accuracy analytics
# - Keep Admin tab visible, cover data with a large in-tab auth overlay
# - Relock Admin tab every time user leaves and re-enters it (Behavior B)
# =========================================================

# -----------------------------
# Machine suggestion / JDD helpers
# -----------------------------
V34_PS4_CASHOUT_YELLOW_LIMIT = 2
V34_PADDLE_TOP_YELLOW_LIMIT = 2
V34_TEXT_EXTS = {'.log', '.txt', '.csv', '.error', '.saved', ''}
V34_NOISE_RX = re.compile(r"org\.eclipse\.swt\.SWT\.error|^\s*at\s+org\.eclipse\.swt\.|^\s*Exception in thread", re.IGNORECASE)
V34_LEVEL_RX = re.compile(r"\b(WARNING|WARN|ERROR|SEVERE|FATAL)\b", re.IGNORECASE)
V34_DONE_ARROW_RX = re.compile(r"DONE->(WARNING|ERROR|SEVERE|FATAL)", re.IGNORECASE)
V34_EXCLUDE_IO_HEAD_RX = re.compile(r"io[-_ ]?head", re.IGNORECASE)
V34_CYCLE_RX = re.compile(r"\b(CashInTp|CashOutTp)\b", re.IGNORECASE)
V34_VS_TOKEN_RX = re.compile(r"(CRS_SAFE:String2_VS[12]_[A-Z0-9_]+)", re.IGNORECASE)
V34_CPP_RX = re.compile(r"\b([A-Za-z0-9_]+\.cpp)\b")
V34_CODE_RX = re.compile(
    r"\b(TransportSensorExceptions|HwExceptions|SwExceptions|BankNoteReaderExceptions)\b"
    r"\s+(\d{1,4})\s+([A-Z0-9_]{3,})",
    re.IGNORECASE,
)


def _v34_is_probably_text(fp: Path, sample_bytes: int = 4096) -> bool:
    try:
        data = fp.open('rb').read(sample_bytes)
        if not data:
            return True
        printable = sum((b in (9, 10, 13)) or (32 <= b <= 126) for b in data)
        return (printable / max(1, len(data))) >= 0.20
    except Exception:
        return False


def _v34_pick_recent_jdd_logs(log_dir: str, limit: int = 20):
    p = Path(log_dir)
    if not p.is_dir():
        return []
    cand = []
    for f in p.iterdir():
        if f.is_file() and f.name.lower().startswith('jdd_') and f.suffix.lower() == '.log':
            try:
                cand.append((f.stat().st_mtime, f.stat().st_size, f))
            except Exception:
                pass
    cand.sort(key=lambda x: x[0], reverse=True)
    return [str(x[2]) for x in cand[:limit]]


def _v34_detect_devices_from_recent_jdds(log_dir: str, limit: int = 20):
    devices = {}
    sources = {}
    for log_path in _v34_pick_recent_jdd_logs(log_dir, limit=limit):
        try:
            vsmap = extract_vs_from_jdd_log_file(log_path) or {}
        except Exception:
            vsmap = {}
        for vs, pair in vsmap.items():
            if vs not in devices:
                devices[vs] = pair
                sources[vs] = log_path
        if 'VS1' in devices and 'VS2' in devices:
            break
    return devices, sources


def _v34_detect_level(line: str):
    m = V34_DONE_ARROW_RX.search(line)
    if m:
        return m.group(1).upper()
    m = V34_LEVEL_RX.search(line)
    if m:
        return m.group(1).upper()
    return 'UNKNOWN'


def _v34_bucket(level: str):
    if level in ('ERROR', 'SEVERE', 'FATAL'):
        return 'RED'
    if level in ('WARNING', 'WARN'):
        return 'YELLOW'
    return 'UNKNOWN'


def _v34_cycle_from_token(token: str):
    t = token.lower()
    if 'cashout' in t:
        return 'CASHOUT'
    if 'cashin' in t:
        return 'CASHIN'
    return 'UNKNOWN'


def _v34_build_markers(lines):
    markers = []
    for idx, line in enumerate(lines, start=1):
        low = line.lower()
        if 'crs_head' in low:
            continue
        m = V34_CYCLE_RX.search(low)
        if m:
            markers.append((idx, _v34_cycle_from_token(m.group(1))))
    return markers


def _v34_nearest_cycle(line_no, markers, max_backtrack=600):
    best = None
    for marker_line, marker_cycle in markers:
        if marker_line > line_no:
            break
        best = (marker_line, marker_cycle)
    if best is None:
        return 'UNKNOWN'
    marker_line, marker_cycle = best
    return marker_cycle if (line_no - marker_line) <= max_backtrack else 'UNKNOWN'


def _v34_build_machine_summary(log_dir: str):
    events = []
    ps4_yellow_cashout_count = 0
    ps4_red_cashout_count = 0
    paddle_top_yellow_count = 0
    paddle_top_red_count = 0
    files_total = files_scanned = files_skipped_nontext = 0

    p = Path(log_dir)
    if not p.is_dir():
        return {
            'suggested': 'N/A',
            'confidence': 'LOW',
            'reasons': ['JDD log folder not found'],
            'events': [],
            'stats': {},
        }

    for fp in p.rglob('*'):
        if not fp.is_file() or 'log_scan_output' in fp.parts:
            continue
        ext = fp.suffix.lower()
        if not (ext in V34_TEXT_EXTS or (ext == '' and '' in V34_TEXT_EXTS)):
            continue
        files_total += 1
        if fp.suffix.lower() != '.csv' and not _v34_is_probably_text(fp):
            files_skipped_nontext += 1
            continue
        files_scanned += 1
        try:
            rows = fp.read_text(encoding='utf-8', errors='ignore').splitlines()
        except Exception:
            continue
        markers = _v34_build_markers(rows)
        for idx, line in enumerate(rows, start=1):
            low = line.lower()
            if V34_EXCLUDE_IO_HEAD_RX.search(line):
                continue
            if 'crs_head' in low:
                continue
            level = _v34_detect_level(line)
            if level == 'UNKNOWN':
                continue
            if V34_NOISE_RX.search(line):
                continue
            cyc = _v34_nearest_cycle(idx, markers)
            buck = _v34_bucket(level)
            m = V34_VS_TOKEN_RX.search(line)
            if not m:
                continue
            src_m = V34_CPP_RX.search(line)
            code_m = V34_CODE_RX.search(line)
            src = src_m.group(1) if src_m else ''
            code = '%s:%s:%s' % (code_m.group(1), code_m.group(2), code_m.group(3)) if code_m else ''
            token = m.group(1)
            events.append({
                'file': str(fp), 'line_no': idx, 'cycle': cyc, 'level': level, 'bucket': buck,
                'vs_token': token, 'code': code, 'source': src, 'text': line.strip(),
            })
            tu = token.upper()
            if 'VS2_PS4' in tu and cyc == 'CASHOUT':
                if buck == 'YELLOW':
                    ps4_yellow_cashout_count += 1
                elif buck == 'RED':
                    ps4_red_cashout_count += 1
            if 'VS1_PADDLE_TOP' in tu:
                if buck == 'YELLOW':
                    paddle_top_yellow_count += 1
                elif buck == 'RED':
                    paddle_top_red_count += 1

    suggested = 'PASS'
    reasons = []
    if ps4_red_cashout_count > 0:
        suggested = 'REWORK'
        reasons.append(f'VS2_PS4 RED in CASHOUT = {ps4_red_cashout_count}')
    else:
        reasons.append(f'VS2_PS4 YELLOW in CASHOUT = {ps4_yellow_cashout_count} (limit={V34_PS4_CASHOUT_YELLOW_LIMIT})')
        if ps4_yellow_cashout_count > V34_PS4_CASHOUT_YELLOW_LIMIT:
            suggested = 'REWORK'
            reasons.append('VS2_PS4 count exceeded limit')
    if paddle_top_red_count > 0:
        suggested = 'REWORK'
        reasons.append(f'VS1_PADDLE_TOP RED = {paddle_top_red_count}')
    else:
        reasons.append(f'VS1_PADDLE_TOP YELLOW = {paddle_top_yellow_count} (limit={V34_PADDLE_TOP_YELLOW_LIMIT})')
        if paddle_top_yellow_count > V34_PADDLE_TOP_YELLOW_LIMIT:
            suggested = 'REWORK'
            reasons.append('VS1_PADDLE_TOP count exceeded limit')

    if (ps4_red_cashout_count > 0) or (paddle_top_red_count > 0):
        confidence = 'HIGH'
    elif (ps4_yellow_cashout_count > 0) or (paddle_top_yellow_count > 0):
        confidence = 'MEDIUM'
    else:
        confidence = 'LOW'

    return {
        'suggested': suggested,
        'confidence': confidence,
        'reasons': reasons,
        'events': events,
        'stats': {
            'files_total': files_total,
            'files_scanned': files_scanned,
            'files_skipped_nontext': files_skipped_nontext,
            'vs_events_found': len(events),
        }
    }


# -----------------------------
# Database extensions for ML feedback
# -----------------------------
_V34_ORIG_DB_INIT = Database._init
_V34_ORIG_INSERT_SESSION = Database.insert_session

def _v34_db_init(self):
    _V34_ORIG_DB_INIT(self)
    cur = self.conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ml_feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL,
            tester TEXT NOT NULL,
            vs TEXT NOT NULL,
            machine_result TEXT NOT NULL,
            machine_confidence TEXT NOT NULL,
            machine_reasons TEXT NOT NULL,
            final_result TEXT NOT NULL,
            matched INTEGER NOT NULL DEFAULT 0,
            tester_note TEXT DEFAULT '',
            error_events_json TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    """)
    self.conn.commit()


def _v34_insert_session(self, *args, **kwargs):
    sid = _V34_ORIG_INSERT_SESSION(self, *args, **kwargs)
    try:
        self._last_inserted_session_id = sid
    except Exception:
        pass
    return sid


def save_ml_feedback(self, session_id: str, tester: str, vs: str,
                     machine_result: str, machine_confidence: str, machine_reasons: str,
                     final_result: str, tester_note: str = '', error_events_json: str = ''):
    matched = 1 if (machine_result or '').strip().upper() == (final_result or '').strip().upper() else 0
    self.conn.execute(
        """INSERT INTO ml_feedback
           (session_id, tester, vs, machine_result, machine_confidence, machine_reasons,
            final_result, matched, tester_note, error_events_json, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (session_id, tester, vs, machine_result, machine_confidence, machine_reasons,
         final_result, matched, tester_note, error_events_json, now_ts())
    )
    self.conn.commit()
    return matched


def accuracy_snapshot(self):
    cur = self.conn.cursor()
    queries = {
        'today': "date(created_at)=date('now','localtime')",
        'week': "date(created_at)>=date('now','localtime','-6 day')",
        'month': "strftime('%Y-%m', created_at)=strftime('%Y-%m','now','localtime')",
        'prev_month': "strftime('%Y-%m', created_at)=strftime('%Y-%m','now','localtime','-1 month')",
    }
    out = {}
    for key, cond in queries.items():
        cur.execute(f"SELECT COUNT(*), COALESCE(SUM(matched),0) FROM ml_feedback WHERE {cond}")
        total, matched = cur.fetchone()
        total = int(total or 0)
        matched = int(matched or 0)
        out[key] = {
            'total': total,
            'matched': matched,
            'conflicts': total - matched,
            'accuracy_pct': round((matched / total) * 100.0, 2) if total else None,
        }
    return out

Database._init = _v34_db_init
Database.insert_session = _v34_insert_session
Database.save_ml_feedback = save_ml_feedback
Database.accuracy_snapshot = accuracy_snapshot


# -----------------------------
# App extensions (single clean patch layer on top of clean V26)
# -----------------------------
_V34_ORIG_INIT_RUNTIME = DNTesterStationApp._init_runtime_state
_V34_ORIG_BUILD_UI = DNTesterStationApp._build_ui
_V34_ORIG_START_TEST = DNTesterStationApp.start_test
_V34_ORIG_FINISH_TEST = DNTesterStationApp.finish_test
_V34_ORIG_SET_RESULT_BUTTONS = DNTesterStationApp._set_result_buttons


def _v34_init_runtime_state(self):
    _V34_ORIG_INIT_RUNTIME(self)
    self.connected_devices_var = tk.StringVar(value='No VS device detected yet')
    self.machine_suggest_var = tk.StringVar(value='N/A')
    self.machine_confidence_var = tk.StringVar(value='N/A')
    self.machine_reason_var = tk.StringVar(value='No analysis yet')
    self.current_machine_summary = {'suggested': 'N/A', 'confidence': 'N/A', 'reasons': [], 'events': []}
    self.current_vs_events = []
    self.machine_note_text = None
    self.detected_vs_map = {}
    self.device_popup = None
    self.connected_devices_window = None
    self.current_source_var = tk.StringVar(value='-')
    self.current_test_var = tk.StringVar(value='Not started')
    self._admin_tab_locked = True
    self.admin_tab_overlay = None
    self.admin_auth_password_var = None
    self.admin_auth_msg_var = None
    self.admin_auth_entry = None
    self.last_match_status_var = tk.StringVar(value='No saved decision yet')


def _v34_find_left_content(self):
    try:
        body = self.winfo_children()[1]
        left = body.winfo_children()[0]
        return left.winfo_children()[0]
    except Exception:
        return None


def _v34_walk_widgets(root):
    try:
        for child in root.winfo_children():
            yield child
            for sub in _v34_walk_widgets(child):
                yield sub
    except Exception:
        return


def _v34_hide_admin_button(self):
    try:
        for w in _v34_walk_widgets(self):
            try:
                if isinstance(w, (tk.Button, ttk.Button)) and str(w.cget('text')) == 'Admin Panel':
                    w.pack_forget()
            except Exception:
                pass
    except Exception:
        pass


def _v34_refresh_devices(self):
    if not getattr(self, 'device_cards_frame', None):
        return
    for w in list(self.device_cards_frame.winfo_children()):
        w.destroy()
    devices, sources = _v34_detect_devices_from_recent_jdds(JDD_LOG_DIR, limit=20)
    self.detected_vs_map = devices or {}
    if not devices:
        tk.Label(self.device_cards_frame, text='No VS device detected in recent JDD logs', bg=COLORS['card'], fg=COLORS['gray'], font=('Segoe UI', 10, 'italic')).pack(anchor='w')
        self.connected_devices_var.set('No VS device detected')
        self.status_var.set('Refresh Devices: no VS device detected in recent JDD logs.')
        return
    parts = []
    for vs in ('VS1', 'VS2'):
        if vs not in devices:
            continue
        mat, sn = devices[vs]
        src = sources.get(vs, '')
        parts.append(f'{vs}: {mat} / {sn}')
        card = tk.Frame(self.device_cards_frame, bg='#f8fafc', bd=1, relief='solid')
        card.pack(fill='x', pady=(0, 8))
        tk.Label(card, text=f'{vs} • recent JDD device', bg='#f8fafc', fg=COLORS['header'], font=('Segoe UI', 11, 'bold')).pack(anchor='w', padx=10, pady=(8, 2))
        tk.Label(card, text=f'Material: {mat}', bg='#f8fafc', fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=10)
        tk.Label(card, text=f'Serial: {sn}', bg='#f8fafc', fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=10)
        tk.Label(card, text=f'Source: {Path(src).name if src else "-"}', bg='#f8fafc', fg=COLORS['gray'], font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(0, 8))
        tk.Button(card, text=f'Test {vs}', bg=COLORS['accent'], fg='white', relief='flat', font=('Segoe UI', 10, 'bold'), command=lambda v=vs: self.start_test(v)).pack(anchor='e', padx=10, pady=(0, 10))
    missing = [v for v in ('VS1', 'VS2') if v not in devices]
    if missing:
        self.connected_devices_var.set(f'{" | ".join(parts)} | Missing: {", ".join(missing)}')
        self.status_var.set(f'Refresh Devices: found {", ".join(sorted(devices.keys()))}, missing {", ".join(missing)}.')
    else:
        self.connected_devices_var.set(' | '.join(parts))
        self.status_var.set('Refresh Devices: recent JDD scan completed.')


def _v34_show_connected_devices_window(self):
    if self.connected_devices_window and self.connected_devices_window.winfo_exists():
        self.connected_devices_window.lift()
        return

    self.connected_devices_window = tk.Toplevel(self)
    self.connected_devices_window.title('Connected VS Devices')
    self.connected_devices_window.geometry('420x280')
    self.connected_devices_window.configure(bg=COLORS['bg'])

    header = tk.Frame(self.connected_devices_window, bg=COLORS['card'])
    header.pack(fill='x', padx=12, pady=12)
    tk.Label(header, text='Connected VS Devices', bg=COLORS['card'], fg=COLORS['header'], font=('Segoe UI', 14, 'bold')).pack(anchor='w')
    tk.Label(header, text='Quick view of devices detected from recent JDD logs.', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', pady=(4, 0))

    body = tk.Frame(self.connected_devices_window, bg=COLORS['bg'])
    body.pack(fill='both', expand=True, padx=12, pady=(0, 12))

    if not self.detected_vs_map:
        tk.Label(body, text='No VS devices found yet. Use Refresh Devices to scan recent logs.', bg=COLORS['bg'], fg=COLORS['gray'], font=('Segoe UI', 10), wraplength=380, justify='left').pack(anchor='w', pady=10)
    else:
        for vs in ('VS1', 'VS2'):
            if vs in self.detected_vs_map:
                mat, sn = self.detected_vs_map[vs]
                frame = tk.Frame(body, bg='#f8fafc', bd=1, relief='solid')
                frame.pack(fill='x', pady=(0, 10))
                tk.Label(frame, text=vs, bg='#f8fafc', fg=COLORS['header'], font=('Segoe UI', 11, 'bold')).pack(anchor='w', padx=10, pady=(8, 2))
                tk.Label(frame, text=f'Material: {mat}', bg='#f8fafc', fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=10)
                tk.Label(frame, text=f'Serial: {sn}', bg='#f8fafc', fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=10, pady=(0, 8))
            else:
                tk.Label(body, text=f'{vs} not detected', bg=COLORS['bg'], fg=COLORS['gray'], font=('Segoe UI', 10, 'italic')).pack(anchor='w', pady=(0, 8))

    footer = tk.Frame(self.connected_devices_window, bg=COLORS['bg'])
    footer.pack(fill='x', padx=12, pady=(0, 12))
    tk.Button(footer, text='Refresh Devices', bg=COLORS['accent'], fg='white', relief='flat', command=lambda: [self._v34_refresh_devices(), self._v34_show_connected_devices_window()]).pack(side='left')
    tk.Button(footer, text='Close', bg=COLORS['card'], relief='flat', command=self.connected_devices_window.destroy).pack(side='right')


def _v34_refresh_machine_suggestion(self):
    summary = _v34_build_machine_summary(JDD_LOG_DIR)
    self.current_machine_summary = summary
    self.current_vs_events = summary.get('events', [])
    self.machine_suggest_var.set(summary.get('suggested', 'N/A'))
    self.machine_confidence_var.set(summary.get('confidence', 'N/A'))
    reasons = summary.get('reasons', []) or ['No machine reasons']
    self.machine_reason_var.set(' | '.join(reasons))
    self.status_var.set('Refresh Suggestion: machine analysis updated from recent JDD logs.')


def _v34_show_accuracy_window(self):
    stats = self.db.accuracy_snapshot()
    win = tk.Toplevel(self)
    win.title('Machine Accuracy Analytics')
    win.configure(bg=COLORS['bg'])
    win.geometry('760x430')
    outer = tk.Frame(win, bg=COLORS['bg'])
    outer.pack(fill='both', expand=True, padx=16, pady=16)
    tk.Label(outer, text='Machine Accuracy Analytics', bg=COLORS['bg'], fg=COLORS['header'], font=('Segoe UI', 16, 'bold')).pack(anchor='w', pady=(0, 12))
    for key, title in [('today', 'Today'), ('week', 'Last 7 Days'), ('month', 'Current Month'), ('prev_month', 'Previous Month')]:
        data = stats.get(key, {})
        acc = data.get('accuracy_pct')
        acc_text = 'N/A' if acc is None else f'{acc:.2f}%'
        row = tk.Frame(outer, bg=COLORS['card'], bd=1, relief='solid')
        row.pack(fill='x', pady=5)
        tk.Label(row, text=f"{title}: tested={data.get('total',0)} | matched={data.get('matched',0)} | conflicts={data.get('conflicts',0)} | accuracy={acc_text}", bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 11)).pack(anchor='w', padx=12, pady=9)
    cur = stats.get('month', {})
    prev = stats.get('prev_month', {})
    if cur.get('accuracy_pct') is None or prev.get('accuracy_pct') is None:
        cmp_txt = 'Month-to-month comparison: N/A until both months have reviewed data.'
    else:
        delta = round(cur['accuracy_pct'] - prev['accuracy_pct'], 2)
        sign = '+' if delta >= 0 else ''
        cmp_txt = f'Month-to-month accuracy change: {sign}{delta:.2f}% (current vs previous month)'
    tk.Label(outer, text=cmp_txt, bg=COLORS['bg'], fg=COLORS['header'], font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(12, 0))


def _v34_build_admin_overlay(self):
    if not hasattr(self, 'admin_tab'):
        return
    try:
        if self.admin_tab_overlay is not None and self.admin_tab_overlay.winfo_exists():
            return
    except Exception:
        pass
    overlay = tk.Frame(self.admin_tab, bg='#d8e3ec')
    self.admin_tab_overlay = overlay
    panel = tk.Frame(overlay, bg='white', bd=1, relief='solid', highlightbackground=COLORS['border'], highlightthickness=1)
    panel.place(relx=0.5, rely=0.5, anchor='center', relwidth=0.92, relheight=0.84)
    panel.pack_propagate(False)
    tk.Label(panel, text='Administrator Authorization Required', bg='white', fg=COLORS['header'], font=('Segoe UI', 16, 'bold'), wraplength=520, justify='left').pack(anchor='w', padx=24, pady=(24, 6))
    tk.Label(panel, text='The Admin tab stays visible, but its data is covered until you authorize. Use your admin password to unlock it.', bg='white', fg=COLORS['muted'], wraplength=520, justify='left', font=('Segoe UI', 10)).pack(anchor='w', padx=24)
    tk.Label(panel, text=f'User: {(getattr(self, "tester", "") or "").strip()}', bg='white', fg=COLORS['gray'], font=('Segoe UI', 10, 'bold'), wraplength=520).pack(anchor='w', padx=24, pady=(14, 4))
    self.admin_auth_password_var = tk.StringVar(value='')
    self.admin_auth_msg_var = tk.StringVar(value='')
    self.admin_auth_show_var = tk.BooleanVar(value=False)
    tk.Label(panel, text='Password', bg='white', fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=24, pady=(8, 2))
    self.admin_auth_entry = ttk.Entry(panel, textvariable=self.admin_auth_password_var, width=42, show='*', font=('Segoe UI', 12))
    self.admin_auth_entry.pack(anchor='w', padx=24, pady=(0, 8))
    def _toggle_pwd():
        try:
            self.admin_auth_entry.configure(show='' if self.admin_auth_show_var.get() else '*')
        except Exception:
            pass
    ttk.Checkbutton(panel, text='Show password', variable=self.admin_auth_show_var, command=_toggle_pwd).pack(anchor='w', padx=24, pady=(0, 8))
    tk.Label(panel, textvariable=self.admin_auth_msg_var, bg='white', fg=COLORS['danger'], font=('Segoe UI', 10)).pack(anchor='w', padx=24)
    tk.Label(panel, text='Press Enter or click Authorize to continue.', bg='white', fg=COLORS['muted'], font=('Segoe UI', 9)).pack(anchor='w', padx=24, pady=(4, 10))
    row = tk.Frame(panel, bg='white')
    row.pack(fill='x', padx=24, pady=(10, 0))
    tk.Button(row, text='Authorize', command=self._v34_unlock_admin_tab, bg=COLORS['accent'], fg='white', relief='flat', font=('Segoe UI', 10, 'bold'), padx=12, pady=8).pack(side='left')
    ttk.Button(row, text='Back to Dashboard', command=self._v34_back_from_admin_overlay).pack(side='left', padx=(10, 0))
    try:
        self.admin_auth_entry.bind('<Return>', lambda e: self._v34_unlock_admin_tab())
    except Exception:
        pass


def _v34_show_admin_overlay(self):
    if not hasattr(self, 'admin_tab'):
        return
    self._v34_build_admin_overlay()
    self._admin_tab_locked = True
    try:
        self.admin_auth_password_var.set('')
        self.admin_auth_msg_var.set('')
        self.admin_tab_overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.admin_tab_overlay.lift()
        self.after(80, lambda: self.admin_auth_entry.focus_set())
    except Exception:
        pass


def _v34_unlock_admin_tab(self):
    pw = ''
    try:
        pw = (self.admin_auth_password_var.get() or '')
    except Exception:
        pass
    if not pw:
        try:
            self.admin_auth_msg_var.set('Enter password')
        except Exception:
            pass
        return
    username = (getattr(self, 'tester', '') or '').strip()
    role = self.db.authenticate(username, pw)
    if role != 'admin':
        try:
            self.admin_auth_msg_var.set('Authorization failed')
            self.admin_auth_password_var.set('')
            self.after(50, lambda: self.admin_auth_entry.focus_set())
        except Exception:
            pass
        return
    self._admin_tab_locked = False
    try:
        self.admin_tab_overlay.place_forget()
        self.admin_auth_msg_var.set('')
        self.admin_auth_password_var.set('')
        self.status_var.set('Admin tab unlocked.')
    except Exception:
        pass


def _v34_back_from_admin_overlay(self):
    try:
        self.notebook.select(0)
    except Exception:
        pass


def _v34_on_tab_changed(self, event=None):
    try:
        current = self.notebook.nametowidget(self.notebook.select())
    except Exception:
        return
    if hasattr(self, 'admin_tab') and current == self.admin_tab:
        self._v34_show_admin_overlay()
    else:
        # behavior B: relock every time user leaves admin tab
        self._admin_tab_locked = True
        try:
            if self.admin_tab_overlay is not None and self.admin_tab_overlay.winfo_exists():
                self.admin_tab_overlay.place_forget()
        except Exception:
            pass


def open_admin_panel(self):
    # keep method available, allow login to admin panel for authorized users
    try:
        self.show_admin_tab()
    except Exception:
        pass


def _v34_build_ui(self):
    _V34_ORIG_BUILD_UI(self)
    left = self._v34_find_left_content()
    if left is None:
        return

    # Prominent in-app banner to prove the modern UI is active
    try:
        banner = tk.Label(left, text='MODERN UI ACTIVE — V34', bg=COLORS['accent'], fg='white', font=('Segoe UI', 11, 'bold'), padx=8, pady=6)
        banner.pack(fill='x', pady=(8, 6))
    except Exception:
        pass

    quick = tk.LabelFrame(left, text='Quick Start', bg=COLORS['card'], fg=COLORS['header'], font=('Segoe UI', 10, 'bold'))
    quick.pack(fill='x', pady=(12, 0))
    quick_inner = tk.Frame(quick, bg=COLORS['card'])
    quick_inner.pack(fill='x', padx=10, pady=10)
    tk.Button(quick_inner, text='Start VS1', bg=COLORS['accent'], fg='white', relief='flat', font=('Segoe UI', 11, 'bold'), command=lambda: self.start_test('VS1')).pack(side='left', expand=True, fill='x', padx=(0, 6))
    tk.Button(quick_inner, text='Start VS2', bg=COLORS['accent'], fg='white', relief='flat', font=('Segoe UI', 11, 'bold'), command=lambda: self.start_test('VS2')).pack(side='left', expand=True, fill='x', padx=(6, 6))
    tk.Button(quick_inner, text='Admin Login', bg=COLORS['doa'], fg='white', relief='flat', font=('Segoe UI', 11, 'bold'), command=self.open_admin_panel).pack(side='left', expand=True, fill='x', padx=(6, 0))
    tk.Label(quick, text='Tap a VS button to begin testing, or open Admin access with a password.', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 9), wraplength=420, justify='left').pack(anchor='w', padx=10, pady=(0, 10))

    # Current test context (inline instead of popup)
    ctx = tk.LabelFrame(left, text='Current Test Context', bg=COLORS['card'], fg=COLORS['header'], font=('Segoe UI', 10, 'bold'))
    ctx.pack(fill='x', pady=(12, 0))
    tk.Label(ctx, text='Current VS', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).grid(row=0, column=0, sticky='w', padx=10, pady=(10, 2))
    tk.Label(ctx, textvariable=self.current_test_var, bg=COLORS['card'], fg=COLORS['header'], font=('Segoe UI', 10, 'bold')).grid(row=0, column=1, sticky='w', padx=10, pady=(10, 2))
    tk.Label(ctx, text='Material', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).grid(row=1, column=0, sticky='w', padx=10, pady=2)
    tk.Label(ctx, textvariable=self.material_var, bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).grid(row=1, column=1, sticky='w', padx=10, pady=2)
    tk.Label(ctx, text='Serial', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).grid(row=2, column=0, sticky='w', padx=10, pady=2)
    tk.Label(ctx, textvariable=self.serial_var, bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).grid(row=2, column=1, sticky='w', padx=10, pady=2)
    tk.Label(ctx, text='Source', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).grid(row=3, column=0, sticky='w', padx=10, pady=(2, 2))
    tk.Label(ctx, textvariable=self.current_source_var, bg=COLORS['card'], fg=COLORS['gray'], font=('Segoe UI', 9)).grid(row=3, column=1, sticky='w', padx=10, pady=(2, 2))
    tk.Label(ctx, text='Machine match', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).grid(row=4, column=0, sticky='w', padx=10, pady=(2, 10))
    tk.Label(ctx, textvariable=self.last_match_status_var, bg=COLORS['card'], fg=COLORS['header'], font=('Segoe UI', 10, 'bold')).grid(row=4, column=1, sticky='w', padx=10, pady=(2, 10))

    # Device panel
    panel = tk.LabelFrame(left, text='Detected VS Devices (recent JDD scan)', bg=COLORS['card'], fg=COLORS['header'], font=('Segoe UI', 10, 'bold'))
    panel.pack(fill='x', pady=(12, 0))
    self.device_cards_frame = tk.Frame(panel, bg=COLORS['card'])
    self.device_cards_frame.pack(fill='x', padx=10, pady=10)
    prow = tk.Frame(panel, bg=COLORS['card'])
    prow.pack(fill='x', padx=10, pady=(0, 10))
    ttk.Button(prow, text='Refresh Devices', command=self.refresh_devices_panel).pack(side='left')
    ttk.Button(prow, text='Device Summary', command=self._v34_show_connected_devices_window).pack(side='left', padx=(8, 0))
    tk.Label(prow, textvariable=self.connected_devices_var, bg=COLORS['card'], fg=COLORS['gray'], font=('Segoe UI', 9), wraplength=250, justify='left').pack(side='left', padx=(8, 0))

    # Suggestion panel
    sugg = tk.LabelFrame(left, text='Machine Suggestion & Feedback', bg=COLORS['card'], fg=COLORS['header'], font=('Segoe UI', 10, 'bold'))
    sugg.pack(fill='x', pady=(12, 0))
    tk.Label(sugg, text='Suggested Result', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=10, pady=(10, 0))
    tk.Label(sugg, textvariable=self.machine_suggest_var, bg=COLORS['card'], fg=COLORS['header'], font=('Segoe UI', 13, 'bold')).pack(anchor='w', padx=10)
    tk.Label(sugg, text='Confidence', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=10, pady=(8, 0))
    tk.Label(sugg, textvariable=self.machine_confidence_var, bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10, 'bold')).pack(anchor='w', padx=10)
    tk.Label(sugg, text='Reasons', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=10, pady=(8, 0))
    tk.Label(sugg, textvariable=self.machine_reason_var, bg=COLORS['card'], fg=COLORS['gray'], wraplength=340, justify='left', font=('Segoe UI', 9)).pack(anchor='w', padx=10)
    tk.Label(sugg, text='Talk to machine (saved with session)', bg=COLORS['card'], fg=COLORS['muted'], font=('Segoe UI', 10)).pack(anchor='w', padx=10, pady=(10, 0))
    self.machine_note_text = tk.Text(sugg, height=4, width=36, wrap='word', font=('Segoe UI', 9))
    self.machine_note_text.pack(fill='x', padx=10, pady=(4, 8))
    srow = tk.Frame(sugg, bg=COLORS['card'])
    srow.pack(fill='x', padx=10, pady=(0, 10))
    ttk.Button(srow, text='Refresh Suggestion', command=self.refresh_machine_suggestion).pack(side='left')
    ttk.Button(srow, text='Accuracy Analytics', command=self.show_accuracy_analytics).pack(side='left', padx=(8, 0))

    self._v34_hide_admin_button()
    try:
        if hasattr(self, 'notebook') and hasattr(self, 'admin_tab'):
            self.notebook.bind('<<NotebookTabChanged>>', self._v34_on_tab_changed)
            self._v34_build_admin_overlay()
    except Exception:
        pass

    self.refresh_devices_panel()
    self.refresh_machine_suggestion()
    self.status_var.set('UI ready. Start a detected VS test, then use PASS / REWORK / DOA.')


def refresh_devices_panel(self):
    self._v34_refresh_devices()


def refresh_machine_suggestion(self):
    self._v34_refresh_machine_suggestion()


def show_accuracy_analytics(self):
    self._v34_show_accuracy_window()


def _v34_set_result_buttons(self, enabled: bool):
    _V34_ORIG_SET_RESULT_BUTTONS(self, enabled)
    # keep original visible buttons exactly as operator controls


def _v34_start_test(self, vs: str):
    # immediate prefill from recent known JDD if available, then continue original log flow
    try:
        self._clear_fields()
    except Exception:
        pass
    self.current_vs = vs
    self.start_time = now_ts()
    self.current_test_var.set(vs)
    if USE_JDD_LOG_EXTRACTION:
        devices, sources = _v34_detect_devices_from_recent_jdds(JDD_LOG_DIR, limit=20)
        if vs in devices:
            mat, sn = devices[vs]
            source_path = sources.get(vs, '')
            self.material_var.set(mat)
            self.serial_var.set(sn)
            self.current_source_var.set(Path(source_path).name if source_path else '-')
            self.current_trc_path = source_path or ''
            self.current_trc_time = now_ts()
            self.status_var.set(f'Preparing {vs} from recent JDD: {mat} / {sn}')
        else:
            self.current_source_var.set('-')
            self.status_var.set(f'Waiting for {vs} in JDD logs...')
        self.refresh_machine_suggestion()
        self._start_vs_from_logs(vs)
        return
    return _V34_ORIG_START_TEST(self, vs)


def _v34_poll_logs(self, vs: str, baseline_mtime: float, started_at: float,
                   last_path: str = '', last_size: int = -1, stable: int = 0):
    if (time.time() - started_at) > JDD_LOG_TIMEOUT_SEC:
        self.status_var.set('Timeout: VS data not found in JDD logs')
        return

    log_path = _jdd_find_new_log(JDD_LOG_DIR, baseline_mtime) if JDD_REQUIRE_NEW_LOG else ''
    if not log_path:
        log_path = _jdd_pick_best_log_for_vs_material(JDD_LOG_DIR, baseline_mtime if JDD_REQUIRE_NEW_LOG else 0.0, target_vs=vs)
    if not log_path:
        self.vs_status_var.set(f'Waiting for JDD logs containing {vs}...')
        self.after(JDD_LOG_POLL_MS, lambda: self._poll_vs_from_logs(vs, baseline_mtime, started_at, last_path, last_size, stable))
        return

    try:
        size = Path(log_path).stat().st_size
    except Exception:
        size = -1
    if log_path == last_path and size == last_size:
        stable += 1
    else:
        stable = 0
    if stable < JDD_LOG_STABLE_PASSES:
        self.vs_status_var.set(f'Checking {Path(log_path).name} for {vs} (stabilizing {stable}/{JDD_LOG_STABLE_PASSES})')
        self.after(JDD_LOG_POLL_MS, lambda: self._poll_vs_from_logs(vs, baseline_mtime, started_at, log_path, size, stable))
        return

    try:
        vsmap = extract_vs_from_jdd_log_file(log_path)
    except Exception:
        vsmap = {}

    if vs in vsmap:
        mat, sn = vsmap[vs]
        self.material_var.set(mat)
        self.serial_var.set(sn)
        self.current_source_var.set(Path(log_path).name)
        self.current_trc_path = log_path
        self.current_trc_time = now_ts()
        self.vs_status_var.set(f'Loaded {vs} from {Path(log_path).name}')
        self.status_var.set(f'(Log) Loaded {vs}: {mat} / {sn} (from {Path(log_path).name})')
        self._set_result_buttons(True)
        self.refresh_devices_panel()
        self.refresh_machine_suggestion()
        return

    found_vs = sorted(vsmap.keys()) if vsmap else []
    if found_vs:
        self.vs_status_var.set(f'Found {", ".join(found_vs)} in {Path(log_path).name}; waiting for {vs}.')
    else:
        self.vs_status_var.set(f'Log {Path(log_path).name} does not contain VS data yet.')
    self.after(JDD_LOG_POLL_MS, lambda: self._poll_vs_from_logs(vs, baseline_mtime, started_at, log_path, size, stable))


def _v34_finish_test(self, result: str):
    prev_sid = getattr(self.db, '_last_inserted_session_id', None)
    machine_result = self.machine_suggest_var.get().strip() if hasattr(self, 'machine_suggest_var') else 'N/A'
    machine_conf = self.machine_confidence_var.get().strip() if hasattr(self, 'machine_confidence_var') else 'N/A'
    machine_reasons = '\n'.join(self.current_machine_summary.get('reasons', [])) if getattr(self, 'current_machine_summary', None) else ''
    note = ''
    try:
        if self.machine_note_text is not None:
            note = self.machine_note_text.get('1.0', 'end').strip()
    except Exception:
        pass

    _V34_ORIG_FINISH_TEST(self, result)
    sid = getattr(self.db, '_last_inserted_session_id', None)
    if sid and sid != prev_sid:
        match_text = 'Match' if machine_result.strip().upper() == result.strip().upper() else 'Mismatch'
        try:
            self.db.save_ml_feedback(
                session_id=sid,
                tester=self.tester or '',
                vs=self.current_vs or '',
                machine_result=machine_result or 'N/A',
                machine_confidence=machine_conf or 'N/A',
                machine_reasons=machine_reasons,
                final_result=result,
                tester_note=note,
                error_events_json=json.dumps(self.current_vs_events, ensure_ascii=False),
            )
            self.last_match_status_var.set(match_text)
        except Exception as exc:
            try:
                messagebox.showwarning('ML feedback', f'Session saved, but ML feedback save failed:\n{exc}')
            except Exception:
                pass
    try:
        if self.machine_note_text is not None:
            self.machine_note_text.delete('1.0', 'end')
    except Exception:
        pass
    try:
        self.current_test_var.set('Not started')
        self.current_source_var.set('-')
        self.refresh_machine_suggestion()
    except Exception:
        pass

DNTesterStationApp._init_runtime_state = _v34_init_runtime_state
DNTesterStationApp._v34_find_left_content = _v34_find_left_content
DNTesterStationApp._v34_hide_admin_button = _v34_hide_admin_button
DNTesterStationApp._v34_refresh_devices = _v34_refresh_devices
DNTesterStationApp._v34_refresh_machine_suggestion = _v34_refresh_machine_suggestion
DNTesterStationApp._v34_show_accuracy_window = _v34_show_accuracy_window
DNTesterStationApp._v34_build_admin_overlay = _v34_build_admin_overlay
DNTesterStationApp._v34_show_admin_overlay = _v34_show_admin_overlay
DNTesterStationApp._v34_unlock_admin_tab = _v34_unlock_admin_tab
DNTesterStationApp._v34_back_from_admin_overlay = _v34_back_from_admin_overlay
DNTesterStationApp._v34_on_tab_changed = _v34_on_tab_changed
DNTesterStationApp._build_ui = _v34_build_ui
DNTesterStationApp.refresh_devices_panel = refresh_devices_panel
DNTesterStationApp.refresh_machine_suggestion = refresh_machine_suggestion
DNTesterStationApp.show_accuracy_analytics = show_accuracy_analytics
DNTesterStationApp._set_result_buttons = _v34_set_result_buttons
DNTesterStationApp.start_test = _v34_start_test
DNTesterStationApp._poll_vs_from_logs = _v34_poll_logs
DNTesterStationApp.finish_test = _v34_finish_test
DNTesterStationApp.open_admin_panel = open_admin_panel

if __name__ == "__main__":
    main()
